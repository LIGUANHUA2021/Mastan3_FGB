# -*- coding: utf-8 -*-

"""
Module implementing ShowResultsMCurv_Dialog.
"""

from PySide6.QtCore import Slot, Qt, QTimer
from PySide6.QtWidgets import QDialog, QHeaderView, QAbstractItemView, QTableWidgetItem
from PySide6 import QtWidgets
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.ticker import ScalarFormatter, FuncFormatter
from .Ui_ShowResultsMCurv import Ui_ShowResultsMCurv_Dialog
from PySide6.QtWidgets import QApplication, QMainWindow, QGraphicsView, QGraphicsScene, QVBoxLayout, QWidget, QSizePolicy
from PySide6.QtGui import QPainter
import numpy as np
import pyqtgraph as pg
from pyqtgraph.Qt import QtGui, QtCore
from openpyxl import Workbook
from analysis.RCD.variables import Model
from analysis.RCD.variables.Model import MomentCurvatureResults as MomCurvaResults


class CustomFigureCanvas(FigureCanvas):
    def __init__(self, fig):
        super().__init__(fig)

    def mouseDoubleClickEvent(self, event):
        if event.button() == Qt.RightButton:
            ax = self.figure.axes[0]
            ax.set_xlim(min(self.figure.axes[0].lines[0].get_xdata() * 1.1), max(self.figure.axes[0].lines[0].get_xdata()) * 1.1)
            ax.set_ylim(min(self.figure.axes[0].lines[0].get_ydata() * 1.1), max(self.figure.axes[0].lines[0].get_ydata()) * 1.1)
            self.draw()

class ShowResultsMCurv_Dialog(QDialog, Ui_ShowResultsMCurv_Dialog):
    """
    Class documentation goes here.
    """

    def __init__(self, parent):
        """
        Constructor

        @param parent reference to the parent widget (defaults to None)
        @type QWidget (optional)
        """
        super().__init__(parent=parent)
        self.setupUi(self)
        self.parent_window = self.parent()
        self.grandparent_window = self.parent_window.parent()
        self.x = np.array([])
        self.y = np.array([])
        #
        self.Curves_tabWidget.setStyleSheet("""
            QTabBar::tab:selected {
                color: white;
            }
            QTabBar::tab:!selected {
                color: black;
            }
        """)
        #---------------------------------------------------------------------
        # fig = Figure(figsize=(5, 4), dpi=100, facecolor="black")
        # #self.canvas = FigureCanvas(fig)
        # self.canvas = CustomFigureCanvas(fig)
        # self.MCurvPlot_verticalLayout.addWidget(self.canvas)
        # self.ax = fig.subplots()
        # ##
        # self.canvas.mpl_connect("scroll_event", self.on_scroll)
        # self.canvas.mpl_connect("button_release_event", self.on_release)
        ##
        # self.ResultsDatasets_tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.ResultsDatasets_tableWidget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        # self.ResultsDatasets_tableWidget.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        # #
        # self.ResultsDatasets_tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
        # self.ResultsDatasets_tableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
        # self.ResultsDatasets_tableWidget.verticalHeader().setVisible(False)
        # self.ResultsDatasets_tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # self.MomCurva_radioButton.toggled.connect(self.MCurvaPlot)
        # self.MomStrn_radioButton.toggled.connect(self.MCurvaPlot)
        # self.MomStrs_radioButton.toggled.connect(self.MCurvaPlot)
        # self.MomTS_radioButton.toggled.connect(self.MCurvaPlot)
        # self.MomSS_radioButton.toggled.connect(self.MCurvaPlot)
        # self.MomCurva_tabWidget.currentChanged.connect(self.MCurvaPlot)
        ##
        #### For tesing
        # self.MCurvaPlot()
        ## Initialize Results Datasets
        self.OMz_y = dict()
        self.Oan_y = dict()
        self.Idn_y = dict()
        self.Stress_y = dict()
        self.Strain_y = dict()
        self.TangentSlope_y = dict()
        self.SecantSlope_y = dict()
        #
        self.OMz_z = dict()
        self.Oan_z = dict()
        self.Idn_z = dict()
        self.Stress_z = dict()
        self.Strain_z = dict()
        self.TangentSlope_z = dict()
        self.SecantSlope_z = dict()
        ##
        self.initDialog()

    def on_scroll(self, event):
        if event.button == "up":
            self.ax.set_xlim(self.ax.get_xlim()[0] * 1.1, self.ax.get_xlim()[1] * 1.1)
            self.ax.set_ylim(self.ax.get_ylim()[0] * 1.1, self.ax.get_ylim()[1] * 1.1)
        elif event.button == "down":
            self.ax.set_xlim(self.ax.get_xlim()[0] * 0.9, self.ax.get_xlim()[1] * 0.9)
            self.ax.set_ylim(self.ax.get_ylim()[0] * 0.9, self.ax.get_ylim()[1] * 0.9)
        self.canvas.draw_idle()

    # def initResultsTable(self):
    #     # self.ResultsDatasets_tableWidget.setRowCount(10)
    #     # # LP2DTable = self.LoadingP_2D_tableWidget()
    #     # for row in range(10):
    #     #     for col in range(self.ResultsDatasets_tableWidget.columnCount()):
    #     #         item = QTableWidgetItem()
    #     #         item.setTextAlignment(Qt.AlignCenter)
    #     #         self.ResultsDatasets_tableWidget.setItem(row, col, item)
    #     #
    #     self.ResultsDatasets_tableWidget.resizeColumnsToContents()
    #     column_labels = ["ID", "Moment", "Curvature", "Depth of Neutral Axis", "Stress", "Strain", "Tangent Slope", "Secant Slope"]
    #     self.ResultsDatasets_tableWidget.setHorizontalHeaderLabels(column_labels)
    #     self.ResultsDatasets_tableWidget.horizontalHeader().setVisible(True)
    #     #
    #     row_count = self.ResultsDatasets_tableWidget.rowCount()
    #     for i in range(row_count):
    #         item = QTableWidgetItem()
    #         item.setText(str(i + 1))
    #         # self.LoadingP_2D_tableWidget.setItem(i + 1, 0, item)
    #         self.ResultsDatasets_tableWidget.setItem(i, 0, item)
    #     #
    #     for row in range(row_count):
    #         for col in range(self.ResultsDatasets_tableWidget.columnCount()):
    #             item = QTableWidgetItem()
    #             item.setTextAlignment(Qt.AlignCenter)
    #             self.ResultsDatasets_tableWidget.setItem(row, col, item)
    #     #
    #     for ii in range(self.ResultsDatasets_tableWidget.rowCount()):
    #         item = self.ResultsDatasets_tableWidget.item(ii, 0)
    #         item.setFlags(item.flags() ^ Qt.ItemIsEditable ^ Qt.ItemIsSelectable)
    #     #
    #     # # Set other column equal widths
    #     self.ResultsDatasets_tableWidget.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
    #     # Set the first and third column adaptive width
    #     self.ResultsDatasets_tableWidget.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
    #     self.ResultsDatasets_tableWidget.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
    #     # Center contents of All cell
    #     for row in range(row_count):
    #         for column in range(row_count):
    #             item = self.ResultsDatasets_tableWidget.item(row, column)
    #             if item is not None:
    #                 item.setTextAlignment(Qt.AlignCenter)

    def initDialog(self):
        # self.initResultsTable()
        self.updateResultTable()
        # self.updateResultTableSetting()
        self.PlotMC()
        #
        self.PlotMStrn()
        #
        self.PlotMStrs()
        #
        self.PlotMTS()
        #
        self.PlotMSS()

        return

    def PlotMC(self):
        for i in reversed(range(self.MCurvPlot_verticalLayout.count())):
            self.MCurvPlot_verticalLayout.itemAt(i).widget().setParent(None)

        if self.parent_window.MyCur_radioButton.isChecked():
            ploteY = MomCurvaResults.OMz_y
            ploteX = MomCurvaResults.Oan_y
            add_key_value = ('Curvature', 0.0)
            ploteY = {add_key_value[0]: add_key_value[1], **ploteY}
            ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
            Xlabel = "Curvature"
        elif self.parent_window.MzCur_radioButton.isChecked():
            ploteY = MomCurvaResults.OMz_z
            ploteX = MomCurvaResults.Oan_z
            add_key_value = ('Curvature', 0.0)
            ploteY = {add_key_value[0]: add_key_value[1], **ploteY}
            ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
            Xlabel = "Curvature"
        ##
        #
        tploteY = ploteY.values()
        tploteX = ploteX.values()
        #
        self.yMC = np.array(list(tploteY))
        self.xMC = np.array(list(tploteX))
        self.plot1 = pg.PlotWidget()
        self.MCurvPlot_verticalLayout.addWidget(self.plot1)
        self.plot1.setBackground('black')
        self.plot1.showGrid(x=True, y=True)
        self.curve1 = self.plot1.plot(self.xMC, self.yMC, pen=pg.mkPen(color='red', width=2))
        Label_style = {'font-size': '12pt', 'font-family': 'Times New Roman'}
        font = QtGui.QFont()
        font.setFamily('Times')
        font.setPointSize(12)
        tAppliedPx = Model.AnalysisInfo.Anap
        if Model.AnalysisInfo.AxialLoadType == 0:
            tempPx = tAppliedPx
        elif Model.AnalysisInfo.AxialLoadType == 1:
            tempPx = str(tAppliedPx) + "%Py"
        self.plot1.setTitle(f"Applied Axial Load Px = {tempPx}", color='w', font = font)
        self.plot1.setLabel("left", "Moment", **Label_style)
        self.plot1.setLabel("bottom", "Curvature", **Label_style)
        self.plot1.showAxis('top')
        self.plot1.showAxis('right')
        self.plot1.getAxis('top').setTicks([])
        self.plot1.getAxis('right').setTicks([])
        self.line = pg.InfiniteLine(angle=90, movable=True, pen='gray', label='')
        self.liney = pg.InfiniteLine(angle=0, movable=True, pen='y', label='')
        self.plot1.addItem(self.line)
        self.plot1.addItem(self.liney)
        self.annotation = pg.TextItem(text='', anchor=(0, 1))
        self.plot1.addItem(self.annotation)

        def on_mouse_move(evt):
            if self.plot1.sceneBoundingRect().contains(evt):
                mouse_point = self.plot1.getViewBox().mapSceneToView(evt)
                x = mouse_point.x()
                if x >= self.xMC[0] and x <= self.xMC[-1]:
                    y_diff = float(np.interp(x, self.xMC, self.yMC))
                    self.line.setValue(x)
                    self.liney.setValue(y_diff)
                    # self.annotation.setText('Moment = {:.3e}'.format(y_diff), color=(128, 128, 128))
                    self.annotation.setText('Moment = {:.3e}\nCurvature = {:.3e}'.format(y_diff, x),
                                            color=(128, 128, 128))
                    self.annotation.setPos(x, y_diff * 0.85)

        self.plot1.scene().sigMouseMoved.connect(on_mouse_move)
        # Set the background color of the PlotWidget to black
        self.plot1.setStyleSheet("background-color: #000000;")

        # Set the color of the axis and labels to white
        styles = {'color': 'w'}
        # set x_axis
        x_axis = self.plot1.getAxis('bottom')
        x_axis.setTextPen(pg.mkPen(**styles))
        x_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
        x_axis.setPen(pg.mkPen(width=2))
        # set y_axis
        y_axis = self.plot1.getAxis('left')
        y_axis.enableAutoSIPrefix(True)
        y_axis.setTextPen(pg.mkPen(**styles))
        y_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
        y_axis.setPen(pg.mkPen(width=2))
        # other axis
        axis = self.plot1.getAxis('right')
        axis.setPen(pg.mkPen(**styles))
        axis.setPen(pg.mkPen(width=2))
        axis = self.plot1.getAxis('top')
        axis.setPen(pg.mkPen(**styles))
        axis.setPen(pg.mkPen(width=2))
        # Legend
        # legend = pg.LegendItem()
        # legend.setGeometry(QtCore.QRectF(0, 0, 150, 100))
        # legend.setParentItem(self.plot1.graphicsItem())
        # legend.addItem(self.curve1, 'Moment Curvature')
        # legend.setAutoFillBackground(True)
        # legend.setBrush('k')
        # legend.anchor((1, 0), (1, 0), offset=(-30, 40))
        return

    def PlotMStrn(self):
        for i in reversed(range(self.MStrnPlot_verticalLayout.count())):
            self.MStrnPlot_verticalLayout.itemAt(i).widget().setParent(None)

        ##
        if self.parent_window.MyCur_radioButton.isChecked():
            ploteY = MomCurvaResults.OMz_y
            ploteX = MomCurvaResults.OutStrn_y
            add_key_value = ('Strain', 0.0)
            ploteY = {add_key_value[0]: add_key_value[1], **ploteY}
            ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
            Xlabel = "Strain"
        elif self.parent_window.MzCur_radioButton.isChecked():
            ploteY = MomCurvaResults.OMz_z
            ploteX = MomCurvaResults.OutStrn_z
            add_key_value = ('Strain', 0.0)
            ploteY = {add_key_value[0]: add_key_value[1], **ploteY}
            ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
            Xlabel = "Strain"
        ##
        #
        tploteY = ploteY.values()
        tploteX = ploteX.values()
        #
        self.yMStrn = np.array(list(tploteY))
        self.xMStrn = np.array(list(tploteX))
        ##
        # self.axMC.clear()
        self.plot2 = pg.PlotWidget()
        self.MStrnPlot_verticalLayout.addWidget(self.plot2)
        self.plot2.setBackground('black')
        self.plot2.showGrid(x=True, y=True)
        self.curve2 = self.plot2.plot(self.xMStrn, self.yMStrn, pen=pg.mkPen(color='red', width=2))
        Label_style = {'font-size': '12pt', 'font-family': 'Times New Roman'}
        font = QtGui.QFont()
        font.setFamily('Times')
        font.setPointSize(12)
        tAppliedPx = Model.AnalysisInfo.Anap
        if Model.AnalysisInfo.AxialLoadType == 0:
            tempPx = tAppliedPx
        elif Model.AnalysisInfo.AxialLoadType == 1:
            tempPx = str(tAppliedPx) + "%Py"
        self.plot2.setTitle(f"Applied Axial Load Px = {tempPx}", color='w', font=font)
        self.plot2.setLabel("left", "Moment", **Label_style)
        self.plot2.setLabel("bottom", Xlabel, **Label_style)
        self.plot2.showAxis('top')
        self.plot2.showAxis('right')
        self.plot2.getAxis('top').setTicks([])
        self.plot2.getAxis('right').setTicks([])
        self.line2 = pg.InfiniteLine(angle=90, movable=True, pen='gray', label='')
        self.liney2 = pg.InfiniteLine(angle=0, movable=True, pen='y', label='')
        self.plot2.addItem(self.line2)
        self.plot2.addItem(self.liney2)
        self.annotation2 = pg.TextItem(text='', anchor=(0, 1))
        self.plot2.addItem(self.annotation2)

        def on_mouse_move2(evt):
            if self.plot2.sceneBoundingRect().contains(evt):
                mouse_point = self.plot2.getViewBox().mapSceneToView(evt)
                x = mouse_point.x()
                if x >= self.xMStrn[0] and x <= self.xMStrn[-1]:
                    y_diff2 = float(np.interp(x, self.xMStrn, self.yMStrn))
                    self.line2.setValue(x)
                    self.liney2.setValue(y_diff2)
                    self.annotation2.setText('Moment = {:.3e}\nStrain = {:.3e}'.format(y_diff2, x),
                                            color=(128, 128, 128))
                    self.annotation2.setPos(x, y_diff2 * 0.85)

        self.plot2.scene().sigMouseMoved.connect(on_mouse_move2)
        # Set the background color of the PlotWidget to black
        self.plot2.setStyleSheet("background-color: #000000;")

        # Set the color of the axis and labels to white
        styles = {'color': 'w'}
        # set x_axis
        x_axis = self.plot2.getAxis('bottom')
        x_axis.setTextPen(pg.mkPen(**styles))
        x_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
        x_axis.setPen(pg.mkPen(width=2))
        # set y_axis
        y_axis = self.plot2.getAxis('left')
        y_axis.enableAutoSIPrefix(True)
        y_axis.setTextPen(pg.mkPen(**styles))
        y_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
        y_axis.setPen(pg.mkPen(width=2))
        # other axis
        axis = self.plot2.getAxis('right')
        axis.setPen(pg.mkPen(**styles))
        axis.setPen(pg.mkPen(width=2))
        axis = self.plot2.getAxis('top')
        axis.setPen(pg.mkPen(**styles))
        axis.setPen(pg.mkPen(width=2))
        # Legend
        # legend = pg.LegendItem()
        # legend.setGeometry(QtCore.QRectF(0, 0, 150, 100))
        # legend.setParentItem(self.plot2.graphicsItem())
        # legend.addItem(self.curve1, 'Moment Curvature')
        # legend.setAutoFillBackground(True)
        # legend.setBrush('k')
        # legend.anchor((1, 0), (1, 0), offset=(-30, 40))
        return

    def PlotMStrs(self):
        for i in reversed(range(self.MStrsPlot_verticalLayout.count())):
            self.MStrsPlot_verticalLayout.itemAt(i).widget().setParent(None)

        ##
        if self.parent_window.MyCur_radioButton.isChecked():
            ploteY = MomCurvaResults.OMz_y
            ploteX = MomCurvaResults.OutStrs_y
            add_key_value = ('Stress', 0.0)
            ploteY = {add_key_value[0]: add_key_value[1], **ploteY}
            ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
            Xlabel = "Stress"
        elif self.parent_window.MzCur_radioButton.isChecked():
            ploteY = MomCurvaResults.OMz_z
            ploteX = MomCurvaResults.OutStrs_z
            add_key_value = ('Stress', 0.0)
            ploteY = {add_key_value[0]: add_key_value[1], **ploteY}
            ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
            Xlabel = "Stress"
        ##
        #
        tploteY = ploteY.values()
        tploteX = ploteX.values()
        #
        self.yMStrs = np.array(list(tploteY))
        self.xMStrs = np.array(list(tploteX))
        ##
        # self.axMC.clear()
        self.plot3 = pg.PlotWidget()
        self.MStrsPlot_verticalLayout.addWidget(self.plot3)
        self.plot3.setBackground('black')
        self.plot3.showGrid(x=True, y=True)
        self.curve3 = self.plot3.plot(self.xMStrs, self.yMStrs, pen=pg.mkPen(color='red', width=2))
        Label_style = {'font-size': '12pt', 'font-family': 'Times New Roman'}
        font = QtGui.QFont()
        font.setFamily('Times')
        font.setPointSize(12)
        ##
        tAppliedPx = Model.AnalysisInfo.Anap
        if Model.AnalysisInfo.AxialLoadType == 0:
            tempPx = tAppliedPx
        elif Model.AnalysisInfo.AxialLoadType == 1:
            tempPx = str(tAppliedPx) + "%Py"

        self.plot3.setTitle(f"Applied Axial Load Px = {tempPx}", color='w', font=font)
        self.plot3.setLabel("left", "Moment", **Label_style)
        self.plot3.setLabel("bottom", Xlabel, **Label_style)
        self.plot3.showAxis('top')
        self.plot3.showAxis('right')
        self.plot3.getAxis('top').setTicks([])
        self.plot3.getAxis('right').setTicks([])
        self.line3 = pg.InfiniteLine(angle=90, movable=True, pen='gray', label='')
        self.liney3 = pg.InfiniteLine(angle=0, movable=True, pen='y', label='')
        self.plot3.addItem(self.line3)
        self.plot3.addItem(self.liney3)
        self.annotation3 = pg.TextItem(text='', anchor=(0, 1))
        self.plot3.addItem(self.annotation3)

        def on_mouse_move3(evt):
            if self.plot3.sceneBoundingRect().contains(evt):
                mouse_point = self.plot3.getViewBox().mapSceneToView(evt)
                x = mouse_point.x()
                if x >= self.xMStrs[0] and x <= self.xMStrs[-1]:
                    y_diff3 = float(np.interp(x, self.xMStrs, self.yMStrs))
                    self.line3.setValue(x)
                    self.liney3.setValue(y_diff3)
                    self.annotation3.setText('Moment = {:.3e}\nStress = {:.3e}'.format(y_diff3, x),
                                            color=(128, 128, 128))
                    self.annotation3.setPos(x, y_diff3)

        self.plot3.scene().sigMouseMoved.connect(on_mouse_move3)
        # Set the background color of the PlotWidget to black
        self.plot3.setStyleSheet("background-color: #000000;")

        # Set the color of the axis and labels to white
        styles = {'color': 'w'}
        # set x_axis
        x_axis = self.plot3.getAxis('bottom')
        x_axis.setTextPen(pg.mkPen(**styles))
        x_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
        x_axis.setPen(pg.mkPen(width=2))
        # set y_axis
        y_axis = self.plot3.getAxis('left')
        y_axis.enableAutoSIPrefix(True)
        y_axis.setTextPen(pg.mkPen(**styles))
        y_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
        y_axis.setPen(pg.mkPen(width=2))
        # other axis
        axis = self.plot3.getAxis('right')
        axis.setPen(pg.mkPen(**styles))
        axis.setPen(pg.mkPen(width=2))
        axis = self.plot3.getAxis('top')
        axis.setPen(pg.mkPen(**styles))
        axis.setPen(pg.mkPen(width=2))
        # Legend
        # legend = pg.LegendItem()
        # legend.setGeometry(QtCore.QRectF(0, 0, 150, 100))
        # legend.setParentItem(self.plot3.graphicsItem())
        # legend.addItem(self.curve1, 'Moment Curvature')
        # legend.setAutoFillBackground(True)
        # legend.setBrush('k')
        # legend.anchor((1, 0), (1, 0), offset=(-30, 40))
        return

    def PlotMTS(self):
        for i in reversed(range(self.MTSPlot_verticalLayout.count())):
            self.MTSPlot_verticalLayout.itemAt(i).widget().setParent(None)
        ##
        ##
        if self.parent_window.MyCur_radioButton.isChecked():
            tploteY = MomCurvaResults.OMz_y
            ttploteY = list(tploteY.items())
            ttploteY.pop()
            ploteY = dict(ttploteY)
            tTS = dict()
            # tOan_y = MomCurvaResults.Oan_y
            tOMz_y_list = list(MomCurvaResults.OMz_y.values())
            tOan_y_list = list(MomCurvaResults.Oan_y.values())
            # print("len(tOMz_y_list)=", len(tOMz_y_list))
            # print("range(len(tOMz_y_list)=", range(len(tOMz_y_list)))
            for ii in range(len(tOMz_y_list)):
                if ii == len(tOMz_y_list) - 1:
                    pass
                else:
                    teTS = (tOMz_y_list[ii + 1] - tOMz_y_list[ii]) / (tOan_y_list[ii + 1] - tOan_y_list[ii])
                    tTS.setdefault(ii, teTS)
            ##
            ploteX = tTS
            add_key_value = ('Tangent_Slope', tTS[0])
            add_key_value2 = ('Tangent_Slope', 0.0)
            ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
            ploteY = {add_key_value2[0]: add_key_value2[1], **ploteY}
            Xlabel = "Tangent Slope"

        elif self.parent_window.MzCur_radioButton.isChecked():
            tploteY = MomCurvaResults.OMz_z
            ttploteY = list(tploteY.items())
            ttploteY.pop()
            ploteY = dict(ttploteY)
            tTS = dict()
            # tOan_z = MomCurvaResults.Oan_z
            tOMz_z_list = list(MomCurvaResults.OMz_z.values())
            tOan_z_list = list(MomCurvaResults.Oan_z.values())
            for ii in range(len(tOMz_z_list)):
                if ii == len(tOMz_z_list) - 1:
                    pass
                else:
                    teTS = (tOMz_z_list[ii + 1] - tOMz_z_list[ii]) / (tOan_z_list[ii + 1] - tOan_z_list[ii])
                    tTS.setdefault(ii, teTS)
            ##
            ploteX = tTS
            add_key_value = ('Tangent_Slope', tTS[0])
            add_key_value2 = ('Tangent_Slope', 0.0)
            ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
            ploteY = {add_key_value2[0]: add_key_value2[1], **ploteY}
            Xlabel = "Tangent Slope"
        ##
        #
        tploteY = ploteY.values()
        tploteX = ploteX.values()
        #
        self.yMTS = np.array(list(tploteY))
        self.xMTS = np.array(list(tploteX))
        ##
        # self.axMC.clear()
        self.plot4 = pg.PlotWidget()
        self.MTSPlot_verticalLayout.addWidget(self.plot4)
        self.plot4.setBackground('black')
        self.plot4.showGrid(x=True, y=True)
        sorted_y = list(self.yMTS)
        sorted_x = list(self.xMTS)
        if 0 in sorted_x:
            index_x = sorted_x.index(0)
        else:
            index_x = len(sorted_x) - 1
        sorted_x = sorted_x[:index_x + 1]
        sorted_y = sorted_y[:index_x + 1]
        sorted_y.reverse()
        sorted_x.reverse()
        self.curve4 = self.plot4.plot(sorted_x, sorted_y, pen=pg.mkPen(color='red', width=2))
        Label_style = {'font-size': '12pt', 'font-family': 'Times New Roman'}
        font = QtGui.QFont()
        font.setFamily('Times')
        font.setPointSize(12)
        ##
        tAppliedPx = Model.AnalysisInfo.Anap
        if Model.AnalysisInfo.AxialLoadType == 0:
            tempPx = tAppliedPx
        elif Model.AnalysisInfo.AxialLoadType == 1:
            tempPx = str(tAppliedPx) + "%Py"
        #
        self.plot4.setTitle(f"Applied Axial Load Px = {tempPx}", color='w', font=font)
        self.plot4.setLabel("left", "Moment", **Label_style)
        self.plot4.setLabel("bottom", Xlabel, **Label_style)
        self.plot4.showAxis('top')
        self.plot4.showAxis('right')
        self.plot4.getAxis('top').setTicks([])
        self.plot4.getAxis('right').setTicks([])
        self.line4 = pg.InfiniteLine(angle=90, movable=True, pen='gray', label='')
        self.liney4 = pg.InfiniteLine(angle=0, movable=True, pen='y', label='')
        self.plot4.addItem(self.line4)
        self.plot4.addItem(self.liney4)
        self.annotation4 = pg.TextItem(text='', anchor=(0, 1))
        self.plot4.addItem(self.annotation4)


        def on_mouse_move4(evt):
            if self.plot4.sceneBoundingRect().contains(evt):
                mouse_point = self.plot4.getViewBox().mapSceneToView(evt)
                x = mouse_point.x()
                # print(x4)
                if x <= sorted_x[-1] and x >= sorted_x[0]:
                    y_diff4 = float(np.interp(float(x), sorted_x, sorted_y))
                    self.line4.setValue(x)
                    self.liney4.setValue(y_diff4)
                    self.annotation4.setText('Moment = {:.3e}\nTangent_Slope = {:.3e}'.format(y_diff4, x),
                                            color=(128, 128, 128))
                    self.annotation4.setPos(x, y_diff4)

        self.plot4.scene().sigMouseMoved.connect(on_mouse_move4)
        # Set the background color of the PlotWidget to black
        self.plot4.setStyleSheet("background-color: #000000;")
        # Set the color of the axis and labels to white
        styles = {'color': 'w'}
        # set x_axis
        x_axis = self.plot4.getAxis('bottom')
        x_axis.setTextPen(pg.mkPen(**styles))
        x_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
        x_axis.setPen(pg.mkPen(width=2))
        # set y_axis
        y_axis = self.plot4.getAxis('left')
        y_axis.enableAutoSIPrefix(True)
        y_axis.setTextPen(pg.mkPen(**styles))
        y_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
        y_axis.setPen(pg.mkPen(width=2))
        # other axis
        axis = self.plot4.getAxis('right')
        axis.setPen(pg.mkPen(**styles))
        axis.setPen(pg.mkPen(width=2))
        axis = self.plot4.getAxis('top')
        axis.setPen(pg.mkPen(**styles))
        axis.setPen(pg.mkPen(width=2))
        # Legend
        # legend = pg.LegendItem()
        # legend.setGeometry(QtCore.QRectF(0, 0, 150, 100))
        # legend.setParentItem(self.plot4.graphicsItem())
        # legend.addItem(self.curve1, 'Moment Curvature')
        # legend.setAutoFillBackground(True)
        # legend.setBrush('k')
        # legend.anchor((1, 0), (1, 0), offset=(-30, 40))
        return

    def PlotMSS(self):
        for i in reversed(range(self.MSSPlot_verticalLayout.count())):
            self.MSSPlot_verticalLayout.itemAt(i).widget().setParent(None)
        ##
        if self.parent_window.MyCur_radioButton.isChecked():
            ploteY = MomCurvaResults.OMz_y
            tSS = dict()
            tOan_y = MomCurvaResults.Oan_y
            for key in ploteY.keys():
                # if key == 0:
                #     tSS[key] = 0.0
                # else:
                teSS = ploteY[key] / tOan_y[key]
                tSS.setdefault(key, teSS)
            #
            ploteX = tSS
            add_key_value = ('Tangent_Slope', tSS[0])
            add_key_value2 = ('Tangent_Slope', 0.0)
            ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
            ploteY = {add_key_value2[0]: add_key_value2[1], **ploteY}
            Xlabel = "Secant Slope"

        elif self.parent_window.MzCur_radioButton.isChecked():
            ploteY = MomCurvaResults.OMz_z
            tSS = dict()
            tOan_z = MomCurvaResults.Oan_z
            for key in ploteY.keys():
                # if key == 0:
                #     tSS[key] = 0.0
                # else:
                teSS = ploteY[key] / tOan_z[key]
                tSS.setdefault(key, teSS)
            #
            ploteX = tSS
            add_key_value = ('Tangent_Slope', tSS[0])
            add_key_value2 = ('Tangent_Slope', 0.0)
            ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
            ploteY = {add_key_value2[0]: add_key_value2[1], **ploteY}
            Xlabel = "Secant Slope"
        ##
        #
        tploteY = ploteY.values()
        tploteX = ploteX.values()
        #
        self.yMSS = list(tploteY)
        self.xMSS = list(tploteX)
        ##
        # self.axMC.clear()
        self.plot5 = pg.PlotWidget()
        self.MSSPlot_verticalLayout.addWidget(self.plot5)
        self.plot5.setBackground('black')
        self.plot5.showGrid(x=True, y=True)
        self.xMSS.reverse()
        self.yMSS.reverse()
        self.curve1 = self.plot5.plot(self.xMSS, self.yMSS, pen=pg.mkPen(color='red', width=2))
        Label_style = {'font-size': '12pt', 'font-family': 'Times New Roman'}
        font = QtGui.QFont()
        font.setFamily('Times')
        font.setPointSize(12)
        ##
        tAppliedPx = Model.AnalysisInfo.Anap
        if Model.AnalysisInfo.AxialLoadType == 0:
            tempPx = tAppliedPx
        elif Model.AnalysisInfo.AxialLoadType == 1:
            tempPx = str(tAppliedPx) + "%Py"

        self.plot5.setTitle(f"Applied Axial Load Px = {tempPx}", color='w', font=font)
        self.plot5.setLabel("left", "Moment", **Label_style)
        self.plot5.setLabel("bottom", Xlabel, **Label_style)
        self.plot5.showAxis('top')
        self.plot5.showAxis('right')
        self.plot5.getAxis('top').setTicks([])
        self.plot5.getAxis('right').setTicks([])
        self.line5 = pg.InfiniteLine(angle=90, movable=True, pen='gray', label='')
        self.liney5 = pg.InfiniteLine(angle=0, movable=True, pen='y', label='')
        self.plot5.addItem(self.line5)
        self.plot5.addItem(self.liney5)
        self.annotation5 = pg.TextItem(text='', anchor=(0, 1))
        self.plot5.addItem(self.annotation5)

        def on_mouse_move5(evt):
            if self.plot5.sceneBoundingRect().contains(evt):
                mouse_point = self.plot5.getViewBox().mapSceneToView(evt)
                x = mouse_point.x()
                if x <= self.xMSS[-1] and x >= self.xMSS[0]:
                    y_diff5 = float(np.interp(x, self.xMSS, self.yMSS))
                    self.line5.setValue(x)
                    self.liney5.setValue(y_diff5)
                    self.annotation5.setText('Moment = {:.3e}\nSecant Slope = {:.3e}'.format(y_diff5, x),
                                             color=(128, 128, 128))
                    self.annotation5.setPos(x, y_diff5 * 0.85)

        self.plot5.scene().sigMouseMoved.connect(on_mouse_move5)
        # Set the background color of the PlotWidget to black
        self.plot5.setStyleSheet("background-color: #000000;")

        # Set the color of the axis and labels to white
        styles = {'color': 'w'}
        # set x_axis
        x_axis = self.plot5.getAxis('bottom')
        x_axis.setTextPen(pg.mkPen(**styles))
        x_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
        x_axis.setPen(pg.mkPen(width=2))
        # set y_axis
        y_axis = self.plot5.getAxis('left')
        y_axis.enableAutoSIPrefix(True)
        y_axis.setTextPen(pg.mkPen(**styles))
        y_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
        y_axis.setPen(pg.mkPen(width=2))
        # other axis
        axis = self.plot5.getAxis('right')
        axis.setPen(pg.mkPen(**styles))
        axis.setPen(pg.mkPen(width=2))
        axis = self.plot5.getAxis('top')
        axis.setPen(pg.mkPen(**styles))
        axis.setPen(pg.mkPen(width=2))
        # Legend
        # legend = pg.LegendItem()
        # legend.setGeometry(QtCore.QRectF(0, 0, 150, 100))
        # legend.setParentItem(self.plot5.graphicsItem())
        # legend.addItem(self.curve1, 'Moment Curvature')
        # legend.setAutoFillBackground(True)
        # legend.setBrush('k')
        # legend.anchor((1, 0), (1, 0), offset=(-30, 40))
        return

    def updateResultTable(self):
        if self.parent_window.MyCur_radioButton.isChecked():
            self.OMz_y = MomCurvaResults.OMz_y
            self.Oan_y = MomCurvaResults.Oan_y
            self.Idn_y = MomCurvaResults.Idn_y
            self.Strain_y = MomCurvaResults.OutStrn_y
            self.Stress_y = MomCurvaResults.OutStrs_y
            # self.Stress_y = dict()
            # self.Strain_y = dict()
            # self.TangentSlope_y = dict()
            # self.SecantSlope_y = dict()
            ## Tangent Slope
            tTS = dict()
            # tOan_y = MomCurvaResults.Oan_y
            tOMz_y_list = list(MomCurvaResults.OMz_y.values())
            tOan_y_list = list(MomCurvaResults.Oan_y.values())
            # print("len(tOMz_y_list)=", len(tOMz_y_list))
            # print("range(len(tOMz_y_list)=", range(len(tOMz_y_list)))
            # print("self.OMz_y.keys() = ", self.OMz_y.keys())
            for ii in range(len(tOMz_y_list)):
                if ii == len(tOMz_y_list) - 1:
                    pass
                else:
                    teTS = (tOMz_y_list[ii + 1] - tOMz_y_list[ii]) / (tOan_y_list[ii + 1] - tOan_y_list[ii])
                    tTS.setdefault(ii, teTS)
            ##
            self.TangentSlope_y = tTS
            ## Secant Slope
            tSS = dict()
            # print("self.OMz_y.keys() = ", self.OMz_y.keys())
            for key in self.OMz_y.keys():
                if key == 0:
                    tSS[key] = 0.0
                else:
                    teSS = self.OMz_y[key] / self.Oan_y[key]
                    tSS.setdefault(key, teSS)
            #
            self.SecantSlope_y = tSS
        elif self.parent_window.MzCur_radioButton.isChecked():
            self.OMz_z = MomCurvaResults.OMz_z
            self.Oan_z = MomCurvaResults.Oan_z
            self.Idn_z = MomCurvaResults.Idn_z
            self.Strain_z = MomCurvaResults.OutStrn_z
            self.Stress_z = MomCurvaResults.OutStrs_z
            ## Tangent Slope
            tTS_z = dict()
            # tOan_y = MomCurvaResults.Oan_y
            tOMz_z_list = list(MomCurvaResults.OMz_z.values())
            tOan_z_list = list(MomCurvaResults.Oan_z.values())
            # print("len(tOMz_y_list)=", len(tOMz_z_list))
            # print("range(len(tOMz_y_list)=", range(len(tOMz_z_list)))
            for ii in range(len(tOMz_z_list)):
                if ii == len(tOMz_z_list) - 1:
                    pass
                else:
                    teTS = (tOMz_z_list[ii + 1] - tOMz_z_list[ii]) / (tOan_z_list[ii + 1] - tOan_z_list[ii])
                    tTS_z.setdefault(ii, teTS)
            ##
            self.TangentSlope_z = tTS_z
            ## Secant Slope
            tSS_z = dict()
            for key in self.OMz_z.keys():
                if key == 1:
                    tSS_z[key] = 0.0
                else:
                    teSS = self.OMz_z[key] / self.Oan_z[key]
                    tSS_z.setdefault(key, teSS)
            #
            self.SecantSlope_z = tSS_z
        #########
        # if self.parent_window.MyCur_radioButton.isChecked():
        #     tOMz_y_list = list(MomCurvaResults.OMz_y.values())
        #     self.ResultsDatasets_tableWidget.setRowCount(len(tOMz_y_list))
        #     for ii in range(len(tOMz_y_list)):
        #         teOMz_y = tOMz_y_list[ii]
        #         self.ResultsDatasets_tableWidget.setItem(ii, 1, QTableWidgetItem(str("{:.4e}".format(teOMz_y))))
        #     ##
        #     tOan_y_list = list(MomCurvaResults.Oan_y.values())
        #     for ii in range(len(tOan_y_list)):
        #         teOan_y = tOan_y_list[ii]
        #         self.ResultsDatasets_tableWidget.setItem(ii, 2, QTableWidgetItem(str("{:.4e}".format(teOan_y))))
        #     ##
        #     tIdn_y_list = list(MomCurvaResults.Idn_y.values())
        #     for ii in range(len(tIdn_y_list)):
        #         teIdn_y = tIdn_y_list[ii]
        #         self.ResultsDatasets_tableWidget.setItem(ii, 3, QTableWidgetItem(str("{:.4e}".format(teIdn_y))))
        #     ##
        #     tStrain_y_list = list(self.Strain_y.values())
        #     for ii in range(len(tStrain_y_list)):
        #         teStrain_y = tStrain_y_list[ii]
        #         self.ResultsDatasets_tableWidget.setItem(ii, 5, QTableWidgetItem(str("{:.4e}".format(teStrain_y))))
        #     ##
        #     tTangentSlope_y_list = list(self.TangentSlope_y.values())
        #     for ii in range(len(tTangentSlope_y_list)):
        #         teTangentSlope_y = tTangentSlope_y_list[ii]
        #         self.ResultsDatasets_tableWidget.setItem(ii, 6, QTableWidgetItem(str("{:.4e}".format(teTangentSlope_y))))
        #     ##
        #     tSecantSlope_y_list = list(self.SecantSlope_y.values())
        #     for ii in range(len(tSecantSlope_y_list)):
        #         teSecantSlope_y = tSecantSlope_y_list[ii]
        #         self.ResultsDatasets_tableWidget.setItem(ii, 7, QTableWidgetItem(str("{:.4e}".format(teSecantSlope_y))))
        # elif self.parent_window.MzCur_radioButton.isChecked():
        #     tOMz_z_list = list(MomCurvaResults.OMz_z.values())
        #     self.ResultsDatasets_tableWidget.setRowCount(len(tOMz_z_list))
        #     for ii in range(len(tOMz_z_list)):
        #         teOMz_z = tOMz_z_list[ii]
        #         self.ResultsDatasets_tableWidget.setItem(ii, 1, QTableWidgetItem(str("{:.4e}".format(teOMz_z))))
        #     ##
        #     tOan_z_list = list(MomCurvaResults.Oan_z.values())
        #     for ii in range(len(tOan_z_list)):
        #         teOan_z = tOan_z_list[ii]
        #         self.ResultsDatasets_tableWidget.setItem(ii, 2, QTableWidgetItem(str("{:.4e}".format(teOan_z))))
        #     ##
        #     tIdn_z_list = list(MomCurvaResults.Idn_z.values())
        #     for ii in range(len(tIdn_z_list)):
        #         teIdn_z = tIdn_z_list[ii]
        #         self.ResultsDatasets_tableWidget.setItem(ii, 3, QTableWidgetItem(str("{:.4e}".format(teIdn_z))))
        #     ##
        #     tStrain_z_list = list(self.Strain_z.values())
        #     for ii in range(len(tStrain_z_list)):
        #         teStrain_z = tStrain_z_list[ii]
        #         self.ResultsDatasets_tableWidget.setItem(ii, 5, QTableWidgetItem(str("{:.4e}".format(teStrain_z))))
        #     ##
        #     tTangentSlope_z_list = list(self.TangentSlope_z.values())
        #     for ii in range(len(tTangentSlope_z_list)):
        #         teTangentSlope_z = tTangentSlope_z_list[ii]
        #         self.ResultsDatasets_tableWidget.setItem(ii, 6, QTableWidgetItem(str("{:.4e}".format(teTangentSlope_z))))
        #     ##
        #     tSecantSlope_z_list = list(self.SecantSlope_z.values())
        #     for ii in range(len(tSecantSlope_z_list)):
        #         teSecantSlope_z = tSecantSlope_z_list[ii]
        #         self.ResultsDatasets_tableWidget.setItem(ii, 7, QTableWidgetItem(str("{:.4e}".format(teSecantSlope_z))))

    # def updateResultTableSetting(self):
    #     ##
    #     row_count = self.ResultsDatasets_tableWidget.rowCount()
    #     for i in range(row_count):
    #         item = QTableWidgetItem()
    #         item.setText(str(i + 1))
    #         # self.LoadingP_2D_tableWidget.setItem(i + 1, 0, item)
    #         self.ResultsDatasets_tableWidget.setItem(i, 0, item)
    #     ##
    #     # Center contents of All cell
    #     for row in range(row_count):
    #         for column in range(row_count):
    #             item = self.ResultsDatasets_tableWidget.item(row, column)
    #             if item is not None:
    #                 item.setTextAlignment(Qt.AlignCenter)
    #     ##
    #     return
    def custom_formatter(x, pos):
        if abs(x) < 900:
            return f"{x:.2f}"
        else:
            return f"{x:.2e}"

    def on_release(self, event):
        if event.button == 3:
            self.right_click_count = 0

    def reset_right_click_count(self):
        self.right_click_count = 0

    def two_decimal_formatter(self, x, pos):
        return f"{x:.2f}"

    def MCurvaPlot(self):
        if self.parent_window.MyCur_radioButton.isChecked():
            if self.MomCurva_radioButton.isChecked():
                ploteY = MomCurvaResults.OMz_y
                ploteX = MomCurvaResults.Oan_y
                add_key_value = ('Curvature', 0.0)
                ploteY = {add_key_value[0]: add_key_value[1], **ploteY}
                ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
                Xlabel = "Curvature"
            elif self.MomStrs_radioButton.isChecked():
                ploteY = MomCurvaResults.OMz_y
                ploteX = MomCurvaResults.Oan_y
                add_key_value = ('Stress', 0.0)
                ploteY = {add_key_value[0]: add_key_value[1], **ploteY}
                ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
                Xlabel = "Stress"
            elif self.MomStrn_radioButton.isChecked():
                ploteY = MomCurvaResults.OMz_y
                ploteX = MomCurvaResults.OutStrn_y
                add_key_value = ('Strain', 0.0)
                ploteY = {add_key_value[0]: add_key_value[1], **ploteY}
                ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
                Xlabel = "Strain"
            elif self.MomTS_radioButton.isChecked():
                tploteY = MomCurvaResults.OMz_y
                ttploteY = list(tploteY.items())
                ttploteY.pop()
                ploteY = dict(ttploteY)
                tTS = dict()
                # tOan_y = MomCurvaResults.Oan_y
                tOMz_y_list = list(MomCurvaResults.OMz_y.values())
                tOan_y_list = list(MomCurvaResults.Oan_y.values())
                # print("len(tOMz_y_list)=", len(tOMz_y_list))
                # print("range(len(tOMz_y_list)=", range(len(tOMz_y_list)))
                for ii in range(len(tOMz_y_list)):
                    if ii == len(tOMz_y_list)-1:
                        pass
                    else:
                        teTS = (tOMz_y_list[ii+1]-tOMz_y_list[ii])/(tOan_y_list[ii+1]-tOan_y_list[ii])
                        tTS.setdefault(ii, teTS)
                ##
                ploteX = tTS
                add_key_value = ('Tangent_Slope', tTS[0])
                add_key_value2 = ('Tangent_Slope', 0.0)
                ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
                ploteY = {add_key_value2[0]: add_key_value2[1], **ploteY}
                Xlabel = "Tangent Slope"

            elif self.MomSS_radioButton.isChecked():
                ploteY = MomCurvaResults.OMz_y
                tSS = dict()
                tOan_y = MomCurvaResults.Oan_y
                for key in ploteY.keys():
                    # if key == 0:
                    #     tSS[key] = 0.0
                    # else:
                    teSS = ploteY[key]/tOan_y[key]
                    tSS.setdefault(key, teSS)
                #
                ploteX = tSS
                add_key_value = ('Tangent_Slope', tSS[0])
                add_key_value2 = ('Tangent_Slope', 0.0)
                ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
                ploteY = {add_key_value2[0]: add_key_value2[1], **ploteY}
                Xlabel = "Secant Slope"

        elif self.parent_window.MzCur_radioButton.isChecked():
            if self.MomCurva_radioButton.isChecked():
                ploteY = MomCurvaResults.OMz_z
                ploteX = MomCurvaResults.Oan_z
                add_key_value = ('Curvature', 0.0)
                ploteY = {add_key_value[0]: add_key_value[1], **ploteY}
                ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
                Xlabel = "Curvature"
            elif self.MomStrs_radioButton.isChecked():
                ploteY = MomCurvaResults.OMz_z
                ploteX = MomCurvaResults.Oan_z
                add_key_value = ('Stress', 0.0)
                ploteY = {add_key_value[0]: add_key_value[1], **ploteY}
                ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
                Xlabel = "Stress"
            elif self.MomStrn_radioButton.isChecked():
                ploteY = MomCurvaResults.OMz_z
                ploteX = MomCurvaResults.OutStrn_z
                add_key_value = ('Strain', 0.0)
                ploteY = {add_key_value[0]: add_key_value[1], **ploteY}
                ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
                Xlabel = "Strain"
            elif self.MomTS_radioButton.isChecked():
                tploteY = MomCurvaResults.OMz_z
                ttploteY = list(tploteY.items())
                ttploteY.pop()
                ploteY = dict(ttploteY)
                tTS = dict()
                # tOan_z = MomCurvaResults.Oan_z
                tOMz_z_list = list(MomCurvaResults.OMz_z.values())
                tOan_z_list = list(MomCurvaResults.Oan_z.values())
                for ii in range(len(tOMz_z_list)):
                    if ii == len(tOMz_z_list)-1:
                        pass
                    else:
                        teTS = (tOMz_z_list[ii + 1] - tOMz_z_list[ii]) / (tOan_z_list[ii + 1] - tOan_z_list[ii])
                        tTS.setdefault(ii, teTS)
                ##
                ploteX = tTS
                add_key_value = ('Tangent_Slope', tTS[0])
                add_key_value2 = ('Tangent_Slope', 0.0)
                ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
                ploteY = {add_key_value2[0]: add_key_value2[1], **ploteY}
                Xlabel = "Tangent Slope"

            elif self.MomSS_radioButton.isChecked():
                ploteY = MomCurvaResults.OMz_z
                tSS = dict()
                tOan_z = MomCurvaResults.Oan_z
                for key in ploteY.keys():
                    # if key == 0:
                    #     tSS[key] = 0.0
                    # else:
                    teSS = ploteY[key] / tOan_z[key]
                    tSS.setdefault(key, teSS)
                #
                ploteX = tSS
                add_key_value = ('Tangent_Slope', tSS[0])
                add_key_value2 = ('Tangent_Slope', 0.0)
                ploteX = {add_key_value[0]: add_key_value[1], **ploteX}
                ploteY = {add_key_value2[0]: add_key_value2[1], **ploteY}
                Xlabel = "Secant Slope"
        #
        tploteY = ploteY.values()
        tploteX = ploteX.values()
        #
        self.y = np.array(list(tploteY))
        self.x = np.array(list(tploteX))
        ##
        # self.x = np.linspace(0, 10, 1000)
        # self.y = np.sin(self.x)
        self.ax.clear()
        self.ax.plot(self.x, self.y, color="red")
        # self.ax.set_xlim(auto=True)  # Reset x-axis range
        # self.ax.set_ylim(auto=True)  # Reset y-axis range
        # Set x and y-axis limits to start from 0
        self.ax.set_xlim(left=0)
        self.ax.set_ylim(bottom=0)
        # Set scientific notation format
        self.ax.xaxis.set_major_formatter(plt.FormatStrFormatter('%.2e'))
        self.ax.yaxis.set_major_formatter(plt.FormatStrFormatter('%.2e'))
        self.ax.grid(linestyle=':', linewidth='0.8', color='gray', alpha=0.8)
        self.canvas.draw_idle()
        # Set tick formatter for both x and y-axis
        # self.ax.xaxis.set_major_formatter(FuncFormatter(self.two_decimal_formatter))
        # self.ax.yaxis.set_major_formatter(FuncFormatter(self.two_decimal_formatter))
        # Set tick formatter for y-axis
        # formatter = ScalarFormatter(useMathText=True)
        # formatter.set_scientific(True)
        # formatter.set_powerlimits((-2, 2))
        # self.ax.xaxis.set_major_formatter(formatter)
        # self.ax.yaxis.set_major_formatter(formatter)
        ##
        ##
        font = {'family': 'Times New Roman', 'size': 12, 'weight': 'bold'}
        ##
        self.ax.set_facecolor("black")
        self.ax.tick_params(axis="x", colors="white", labelsize=12)
        self.ax.tick_params(axis="y", colors="white", labelsize=12)
        self.ax.spines['bottom'].set_color('white')
        self.ax.spines['top'].set_color('white')
        self.ax.spines['right'].set_color('white')
        self.ax.spines['left'].set_color('white')
        for label in self.ax.get_xticklabels() + self.ax.get_yticklabels():
            label.set_fontname('Times New Roman')
        ##
        self.ax.tick_params(direction='in')
        self.ax.set_xlabel(Xlabel, color='white', fontsize=12, fontdict=font)
        self.ax.set_ylabel('Moment', color='white', fontsize=12, fontdict=font)
        ##
        tAppliedPx = Model.AnalysisInfo.Anap
        # print("tAppliedPx=", tAppliedPx)
        if Model.AnalysisInfo.AxialLoadType == 0:
            tempPx = tAppliedPx
        elif Model.AnalysisInfo.AxialLoadType == 1:
            tempPx = str(tAppliedPx)+"%Py"
        self.ax.set_title(f"Applied Axial Load Px = {tempPx}", color='white', fontsize=12, fontdict=font)
        #
        print("x min:", self.x.min(), "x max:", self.x.max())
        print("y min:", self.y.min(), "y max:", self.y.max())
        # # Set tick formatter for y-axis
        # formatter = ScalarFormatter(useMathText=True)
        # formatter.set_scientific(True)
        # formatter.set_powerlimits((-2, 2))
        # self.ax.xaxis.set_major_formatter(formatter)
        # self.ax.yaxis.set_major_formatter(formatter)
        #
        self.canvas.draw_idle()
        #
        return

    @Slot()
    def on_ExportMCurv_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError
        # wb = Workbook()
        # ws = wb.active
        # ws.title = "Moment Curvature"
        #
        # # Write the data to the worksheet
        # ws.cell(row=1, column=1, value="Curvature")
        # ws.cell(row=1, column=2, value="Moment of Z")
        # for i in range(len(self.x)):
        #     ws.cell(row=i + 2, column=1, value=self.x[i])
        #     ws.cell(row=i + 2, column=2, value=self.y[i])
        if self.parent_window.MyCur_radioButton.isChecked():
            dict_list = [
                {'Moment Curvature': (self.Oan_y, self.OMz_y, ['Curvature', 'Moment'])},
                {'Moment vs. Strain': (self.Strain_y, self.OMz_y, ['Strain', 'Moment'])},
                {'Moment vs. Stress': (self.Stress_y, self.OMz_y, ['Stress', 'Moment'])},
                {'Moment vs. Idn': (self.Idn_y, self.OMz_y, ['Idn', 'Moment'])},
                {'Moment vs. Tangent Slope': (self.TangentSlope_y, self.OMz_y, ['Tangent Slope', 'Moment'])},
                {'Moment vs. Secant Slope': (self.SecantSlope_y, self.OMz_y, ['Secant Slope', 'Moment'])},
            ]
        elif self.parent_window.MzCur_radioButton.isChecked():
            dict_list = [
                {'Moment Curvature': (self.Oan_z, self.OMz_z, ['Curvature', 'Moment'])},
                {'Moment vs. Strain': (self.Strain_z, self.OMz_z, ['Strain', 'Moment'])},
                {'Moment vs. Stress': (self.Stress_z, self.OMz_z, ['Stress', 'Moment'])},
                {'Moment vs. Idn': (self.Idn_z, self.OMz_z, ['Idn', 'Moment'])},
                {'Moment vs. Tangent Slope': (self.TangentSlope_z, self.OMz_z, ['Tangent Slope', 'Moment'])},
                {'Moment vs. Secant Slope': (self.SecantSlope_z, self.OMz_z, ['Secant Slope', 'Moment'])},
            ]

        # Create an Excel workbook
        wb = Workbook()

        # Remove the default worksheet that is automatically generated
        wb.remove(wb.active)

        # Iterate through the list of dictionaries
        for data in dict_list:
            # Iterate through the keys and values of the dictionary
            for sheet_name, (sheet_data1, sheet_data2, labels) in data.items():
                # Create a new worksheet
                ws = wb.create_sheet(sheet_name)

                # Write the column labels to the first row of the worksheet
                ws.cell(row=1, column=1, value=labels[0])
                ws.cell(row=1, column=2, value=labels[1])

                # Write the values of the first dictionary to the first column of the worksheet
                for row_idx, (_, value) in enumerate(sheet_data1.items(), start=2):
                    ws.cell(row=row_idx, column=1, value=value)

                # Write the values of the second dictionary to the second column of the worksheet
                for row_idx, (_, value) in enumerate(sheet_data2.items(), start=2):
                    ws.cell(row=row_idx, column=2, value=value)
        # Open a dialog to choose the save location
        save_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")

        if save_path:
            # Save the workbook to the chosen location
            wb.save(save_path)

    @Slot()
    def on_Close_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        #raise NotImplementedError
        QDialog.close(self)
