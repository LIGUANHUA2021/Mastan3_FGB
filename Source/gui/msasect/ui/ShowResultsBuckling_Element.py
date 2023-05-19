from PySide6.QtCore import Slot, Signal
from PySide6.QtGui import QIcon
from openpyxl import Workbook
import numpy as np
from gui.msasect.ui.Ui_ShowResultsBuckling_LineElement import Ui_GlobalBucklingPlot_Dialog
from PySide6.QtWidgets import QApplication, QDialog, QVBoxLayout, QWidget, QFileDialog
from gui.msasect.base.Model import GlobalBuckling
import pyqtgraph as pg
from pyqtgraph.Qt import QtGui, QtCore
from gui.msasect.ui.CalGlobalBuckling_Element import MultiThread

class GlobalBuckling_Element_Plot_Dialog(QDialog, Ui_GlobalBucklingPlot_Dialog):
    """
    Class documentation goes here.
    """
    def __init__(self, mw, parent=None):
        """
        Constructor

        @param parent reference to the parent widget (defaults to None)
        @type QWidget (optional)
        """
        super().__init__(parent)
        self.setupUi(self)
        self.mw = mw
        self.setWindowIcon(QIcon('ui/ico/GlobalBuckling.png'))
        self.PlotBucklingCurves_Bending()

    @Slot()
    def PlotBucklingCurves_Bending(self):
        Buckling_data = GlobalBuckling.Buckling_data
        Buckling_data1 = GlobalBuckling.Buckling_data1
        Lamuda = Buckling_data['Lamuda']
        A =  MultiThread.Parameters.Area
        Iv = MultiThread.Parameters.MomentofInertia_v
        Iw = MultiThread.Parameters.MomentofInertia_w
        I = min(Iv, Iw)
        if len(Buckling_data['Factors1']) > 0:
            LF1 = np.array(Buckling_data['Factors1'])
            if Buckling_data1['Material'] == 'Elastic':
                LF1 = self.find_min_elements(LF1, Buckling_data1['Elastic Factor'], [], [])
            elif Buckling_data1['Material'] == 'Plastic':
                LF1 = self.find_min_elements(LF1, Buckling_data1['Plastic Factor'], [], [])
            else:
                LF1 = self.find_min_elements(LF1, [], [], [])
        if len(Buckling_data['Factors2']) > 0:
            LF2 = np.array(Buckling_data['Factors2'])
            if Buckling_data1['Material'] == 'Elastic':
                LF2 = self.find_min_elements(LF2, Buckling_data1['Elastic Factor'], [], [])
            elif Buckling_data1['Material'] == 'Plastic':
                LF2 = self.find_min_elements(LF2, Buckling_data1['Plastic Factor'], [], [])
            else:
                LF2 = self.find_min_elements(LF2, [], [], [])
        self.plot = pg.PlotWidget()
        self.Compression_Layout.addWidget(self.plot)
        self.plot.setBackground('black')
        self.plot.showGrid(x=True, y=True)
        self.curve1 = self.plot.plot(Lamuda, LF1, pen=pg.mkPen(color='red', width=3))
        self.curve2 = self.plot.plot(Lamuda, LF2, pen=pg.mkPen(color='yellow', width=3, style=QtCore.Qt.DashLine))
        self.curvea = self.plot.plot([0, 0.0001], [max(LF1) * 0.98, max(LF1) * 1.02],
                                     pen=pg.mkPen(color='w', width=0.01))
        Label_style = {'font-size': '12pt', 'font-family': 'Times New Roman'}
        font = QtGui.QFont()
        font.setFamily('Times')
        font.setPointSize(12)
        self.plot.setTitle("Global Buckling Analysis Using Line Element", color='w', font=font)
        self.plot.setLabel("left", "Load Factor", **Label_style)
        self.plot.setLabel("bottom", "λ", **Label_style)
        self.plot.showAxis('top')
        self.plot.showAxis('right')
        self.plot.getAxis('top').setTicks([])
        self.plot.getAxis('right').setTicks([])
        self.line = pg.InfiniteLine(angle=90, movable=True, pen='gray', label='')
        self.plot.addItem(self.line)
        self.annotation = pg.TextItem(text='', anchor=(0, 1))
        self.plot.addItem(self.annotation)

        def on_mouse_move(evt):
            if self.plot.sceneBoundingRect().contains(evt):
                mouse_point = self.plot.getViewBox().mapSceneToView(evt)
                x = mouse_point.x()
                if x >= Lamuda[0] and x <= Lamuda[-1]:
                    LF1_data = float(np.interp(x, Lamuda, LF1))
                    LF2_data = float(np.interp(x, Lamuda, LF2))
                    y_diff = np.abs((LF1_data - LF2_data) / LF2_data)
                    self.line.setValue(x)
                    self.annotation.setText('L = {:.3f}\nLF1 = {:.3e}\nLF2 = {:.3e}\nDiff = (LF1-LF2)/LF2 ={:.2f}%'.format(x*np.sqrt(I/A),LF1_data,LF2_data,y_diff * 100), color=(128, 128, 128))
                    self.annotation.setPos(x, 0.9*np.mean([LF1_data, LF2_data]))

        self.plot.scene().sigMouseMoved.connect(on_mouse_move)
        # Set the background color of the PlotWidget to black
        self.plot.setStyleSheet("background-color: #000000;")

        # Set the color of the axis and labels to white
        styles = {'color': 'w'}
        # set x_axis
        x_axis = self.plot.getAxis('bottom')
        x_axis.setTextPen(pg.mkPen(**styles))
        x_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
        x_axis.setPen(pg.mkPen(width=2))
        # set y_axis
        y_axis = self.plot.getAxis('left')
        y_axis.enableAutoSIPrefix(True)
        y_axis.setTextPen(pg.mkPen(**styles))
        y_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
        y_axis.setPen(pg.mkPen(width=2))
        # other axis
        axis = self.plot.getAxis('right')
        axis.setPen(pg.mkPen(**styles))
        axis.setPen(pg.mkPen(width=2))
        axis = self.plot.getAxis('top')
        axis.setPen(pg.mkPen(**styles))
        axis.setPen(pg.mkPen(width=2))
        # Legend
        legend = pg.LegendItem()
        legend.setGeometry(QtCore.QRectF(0, 0, 150, 100))
        legend.setParentItem(self.plot.graphicsItem())
        legend.addItem(self.curve1, 'Considering twisting effects (LF1)')
        legend.addItem(self.curve2, 'Ignoring twisting effects (LF2)')
        legend.setAutoFillBackground(True)
        legend.setBrush('k')
        legend.anchor((1, 0), (1, 0), offset=(-30, 40))

    @Slot()
    def on_ExportResults_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError
        Buckling_data = GlobalBuckling.Buckling_data
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = 'Eigen-buckling analysis'
        # Write the data to the worksheet
        Lamuda = np.array(Buckling_data['Lamuda'])
        sheet1.cell(row=1, column=1, value='λ')
        for i in range(len(Lamuda)):
            sheet1.cell(row=i + 2, column=1, value=Lamuda[i])
        ii = 2
        if len(Buckling_data['Factors1']) > 0:
            LF1 = np.array(Buckling_data['Factors1'])
            sheet1.cell(row=1, column=ii, value='Considering twisting effects')
            for i in range(len(LF1)):
                sheet1.cell(row=i + 2, column=ii, value=LF1[i])
            ii += 1
        if len(Buckling_data['Factors2']) > 0:
            LF2 = np.array(Buckling_data['Factors2'])
            sheet1.cell(row=1, column=ii, value='Ignoring twisting effects')
            for i in range(len(LF2)):
                sheet1.cell(row=i + 2, column=ii, value=LF2[i])
        save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", str("Global Buckling Sect_"),"Excel Files (*.xlsx)")
        if save_path:
            # Save the workbook to the chosen location
            wb.save(save_path)

    @Slot()
    def on_Close_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError
        QDialog.close(self)

    def find_min_elements(self, list1, list2, list3, list4):
        max_len = max(len(list1), len(list2), len(list3), len(list4))
        result = []
        for i in range(max_len):
            val1 = list1[i] if i < len(list1) else float('inf')
            val2 = list2[i] if i < len(list2) else float('inf')
            val3 = list3[i] if i < len(list3) else float('inf')
            val4 = list4[i] if i < len(list4) else float('inf')
            result.append(min(val1, val2, val3, val4))
        return result