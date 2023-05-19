# -*- coding: utf-8 -*-

"""
Module implementing ShowResultsYS2D_Dialog.
"""
from PySide6 import QtWidgets, QtCore
from PySide6.QtCore import Slot, Qt
from PySide6.QtWidgets import QDialog, QHeaderView, QAbstractItemView, QTableWidgetItem, QLineEdit, QStyledItemDelegate
# from .Ui_AnalYieldSurfaces import Ui_YieldSurfaces_Dialog
from gui.msasect.ui.Ui_AnalYieldSurfaces import Ui_YieldSurfaces_Dialog
from openpyxl import Workbook
from .Ui_ShowResultsYieldS2D import Ui_ShowResultsYS2D_Dialog
from gui.msasect.ui import Ui_MainWindow

import sys
import traceback
from PySide6.QtCore import Qt
from PySide6.QtGui import QIcon, QDoubleValidator
from PySide6.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QWidget
import numpy as np
import matplotlib
matplotlib.use('Qt5Agg')
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
from analysis.CMSect.variables.Model import YieldSAnalResults as CMYieldSAnalResults
from analysis.CMSect.variables.Model import YieldSurfaceAnalInfo
from analysis.FESect.variables.Model import YieldSAnalResults as FEYieldSAnalResults
from analysis.FESect.variables.Model import YieldSurfaceAnalInfo
import matplotlib.font_manager as font_manager
##
from analysis.CMSect.util import GetSectionCapacityFactor2D
## Only for testing
#import pandas as pd

class NumericDelegate(QStyledItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.validator = QDoubleValidator()

    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        editor.setValidator(self.validator)
        return editor

class ShowResultsYS2D_Dialog(QDialog, Ui_ShowResultsYS2D_Dialog):
    """
    Class documentation goes here.
    """

    def __init__(self, mw, parent):
        """
        Constructor

        @param parent reference to the parent widget (defaults to None)
        @type QWidget (optional)
        """
        super().__init__(parent=parent)
        self.setupUi(self)
        self.setWindowIcon(QIcon('ui/ico/YieldSurface.ico'))
        ##
        #self.grandparent = parent.parentWidget().parentWidget()
        self.parent_window = self.parent()
        self.grandparent_window = self.parent_window.parent()
        ##
        self.setFixedSize(950, 525)
        ##
        self.mw = mw
        self.PeY = np.array([])
        self.PeX = np.array([])
        self.TableHX = ''
        self.TableHY = ''
        self.SheetName = ''
        ##
        # Create the figure and add to the layout
        self.fig2D = Figure(facecolor="black")
        self.ax = self.fig2D.add_subplot(111)
        self.canvas = FigureCanvas(self.fig2D)
        self.YS2DPlot_verticalLayout.addWidget(self.canvas)
        ##
        self.lines = []
        ##
        self.LoadingP_2D_tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.LoadingP_2D_tableWidget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.LoadingP_2D_tableWidget.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.LoadingP_2D_tableWidget.verticalHeader().setVisible(False)
        self.LoadingP_2D_tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.LoadingP_2D_tableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
        # self.LoadingP_2D_tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.LoadingP_2D_tableWidget.horizontalHeader().setSectionsClickable(False)
        self.LoadingP_2D_tableWidget.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.LoadingP_2D_tableWidget.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        #self.LoadingP_2D_tableWidget.setTextAlignment(Qt.AlignCenter)
        # self.LoadingP_2D_tableWidget.itemChanged.connect(self.LP2DTable_item_changed)
        # Connect cellChanged signal to custom slot
        self.LoadingP_2D_tableWidget.cellChanged.connect(self.on_cell_changed)
        self.LoadingP_2D_tableWidget.cellClicked.connect(self.on_cell_clicked)
        #
        numeric_delegate = NumericDelegate(self.LoadingP_2D_tableWidget)
        self.LoadingP_2D_tableWidget.setItemDelegate(numeric_delegate)
        ##
        # if len(YieldSAnalResults.ONx_y) > 0 and len(YieldSAnalResults.OMy_x) > 0:
        #      self.mw.PYSPMy_radioButton.setCheckable(True)
        #      self.mw.ShowResults_pushButton.setEnabled(True)
        # else:
        #     self.mw.PYSPMy_radioButton.setCheckable(False)
        # if len(YieldSAnalResults.ONx_z) > 0 and len(YieldSAnalResults.OMz_x) > 0:
        #      self.mw.PYSPMz_radioButton.setCheckable(True)
        # else:
        #     self.mw.PYSPMz_radioButton.setCheckable(False)
        # if len(YieldSAnalResults.OMy_z) > 0 and len(YieldSAnalResults.OMz_y) > 0:
        #      self.mw.PYSMyMz_radioButton.setCheckable(True)
        # else:
        #     self.mw.PYSMyMz_radioButton.setCheckable(False)
        self.LPTableData_dict = {}
        ##
        self.initDialog()
        # self.LoadingP_2D_tableWidget.item(1, 1).setFlags(self.LoadingP_2D_tableWidget.item(1, 1).flags() & ~Qt.ItemIsEnabled)

    def initLoadingPTable(self):
        ##
        self.LoadingP_2D_tableWidget.setRowCount(2)
        # LP2DTable = self.LoadingP_2D_tableWidget()
        for row in range(2):
            for col in range(self.LoadingP_2D_tableWidget.columnCount()):
                item = QTableWidgetItem()
                item.setTextAlignment(Qt.AlignCenter)
                self.LoadingP_2D_tableWidget.setItem(row, col, item)
        ##
        self.LoadingP_2D_tableWidget.resizeColumnsToContents()
        # self.LoadingP_2D_tableWidget.setSpan(0, 0, 1, 0)  ## ID
        ##
        if self.parent_window.PYSPMy_radioButton.isChecked():
            column_labels = ["ID", "Px", "Mv", "SCF"]
            # self.LoadingP_2D_tableWidget.setHorizontalHeaderLabels(["ID", "Px", "Mv", "SCF"])
        elif self.parent_window.PYSMyMz_radioButton.isChecked():
            column_labels = ["ID", "Mv", "Mw", "SCF"]
            # self.LoadingP_2D_tableWidget.setHorizontalHeaderLabels(["ID", "Mv", "Mw", "SCF"])
        elif self.parent_window.PYSPMz_radioButton.isChecked():
            column_labels = ["ID", "Px", "Mw", "SCF"]
            # self.LoadingP_2D_tableWidget.setHorizontalHeaderLabels(["ID", "Px", "Mw", "SCF"])
        ##
        # column_labels = ["ID", "Px", "Mv", "Mw", "SCF"]
        self.LoadingP_2D_tableWidget.setHorizontalHeaderLabels(column_labels)
        self.LoadingP_2D_tableWidget.horizontalHeader().setVisible(True)
        # Move the table header text to the cell in the first row and set the text to be centered
        # for i, label in enumerate(column_labels):
        #     item = QTableWidgetItem(label)
        #     item.setTextAlignment(Qt.AlignCenter)
        #     item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Set the non-editable flag
        #     self.LoadingP_2D_tableWidget.setItem(0, i, item)
        # for col in range(self.LoadingP_2D_tableWidget.columnCount()):
        #     item = QTableWidgetItem('Editable Cell')
        #     item.setFlags(item.flags() ^ Qt.ItemIsEditable)
        #     self.LoadingP_2D_tableWidget.setItem(0, col, item)
        ##
        # for col, text in enumerate(['', 'kN', 'kN.m', '']):
        # for col, text in enumerate(['', ' ', ' ', '']):
        #     item = QTableWidgetItem(text)
        #     item.setTextAlignment(Qt.AlignCenter)
        #     item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)  # remove editable flag
        #     self.LoadingP_2D_tableWidget.setItem(0, col, item)
        # self.LoadingP_2D_tableWidget.setItem(0, 1, QTableWidgetItem('kN'))
        # self.LoadingP_2D_tableWidget.setItem(0, 2, QTableWidgetItem('kN.m'))
        # self.LoadingP_2D_tableWidget.setItem(0, 3, QTableWidgetItem('kN.m'))
        # self.LoadingP_2D_tableWidget.setSpan(0, 3, 1, 3)  ## SCF
        ##
        row_count = self.LoadingP_2D_tableWidget.rowCount()
        for i in range(row_count):
            item = QTableWidgetItem()
            item.setText(str(i + 1))
            # self.LoadingP_2D_tableWidget.setItem(i + 1, 0, item)
            self.LoadingP_2D_tableWidget.setItem(i, 0, item)
        ##
        # for j in range(self.LoadingP_2D_tableWidget.columnCount()):
        #     item = self.LoadingP_2D_tableWidget.item(0, j)
        #     item.setFlags(item.flags() ^ Qt.ItemIsEditable ^ Qt.ItemIsSelectable)
        for ii in range(self.LoadingP_2D_tableWidget.rowCount()):
            item = self.LoadingP_2D_tableWidget.item(ii, 0)
            item.setFlags(item.flags() ^ Qt.ItemIsEditable ^ Qt.ItemIsSelectable)
        # for col in range(self.LoadingP_2D_tableWidget.columnCount()):
        #     self.item(1, col).setFlags(self.item(1, col).flags() ^ Qt.ItemIsSelectable)
        # Set the first column adaptive width
        self.LoadingP_2D_tableWidget.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        # Set other column equal widths
        self.LoadingP_2D_tableWidget.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        ## set the first row as non-editable
        # for col in range(self.LoadingP_2D_tableWidget.columnCount()):
        #     item = self.LoadingP_2D_tableWidget.item(1, col)
        #     item.setFlags(item.flags() & ~Qt.ItemIsEditable)

    def initDialog(self):
        ##
        self.YS2DPlot()
        self.initLoadingPTable()
        self.updat_table_status()

        # self.LoadingP_2D_tableWidget.insertRow(row_count)

    def on_cell_changed(self, row, column):
        # if row == 0:
        #     return
        ##
        if column not in [1, 2]:
            return
        ##
        item1 = self.LoadingP_2D_tableWidget.item(row, 1)
        item2 = self.LoadingP_2D_tableWidget.item(row, 2)
        if item1 is not None and item2 is not None:
            try:
                value1 = float(item1.text()) #if item1 is not None else 0
                value2 = float(item2.text()) #if item2 is not None else 0
                #self.ax.plot(value1, value2, 'ro')
                #self.add_point(value2, value1)
                # self.add_line(np.array([0.0, value2]), np.array([0.0, value1]))
                ##
                MomInterCurve = np.concatenate((self.PeX.reshape(-1, 1), self.PeY.reshape(-1, 1)), axis=1)
                if abs(value1) < 1e-3 and abs(value2) < 1e-3:
                    SCF2D = 0.0
                else:
                    SCF2D = GetSectionCapacityFactor2D.GetSCF2D(MomInterCurve, np.array([value2, value1]))
                ##
                if SCF2D > 1.0:
                    tcolor = 'red'
                else:
                    tcolor = 'green'
                self.add_point(row, value2, value1, tcolor)
                item = QTableWidgetItem(str(format(SCF2D, '.2f')))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.LoadingP_2D_tableWidget.setItem(row, 3, item)
                ##
                # print(f"Row {row}, Col 1: {value1}, Col 2: {value2}")
                # print("Col 1:", value1)
                # print("Col 2:", value2)
            except ValueError:
                #traceback.print_exc()
                print()
        ##
        ## Save the Loading points data from GUI
        # for i in range(1, self.LoadingP_2D_tableWidget.rowCount()):
        #     row = []
        #     # for j in range(self.LoadingP_2D_tableWidget.columnCount()):
        #     for j in [0, 1, 2]:  ## Loading points
        #         item = self.LoadingP_2D_tableWidget.item(i, j)
        #         if item is not None:
        #             row.append(item.text())
        #         else:
        #             row.append("")
        #     self.LPTableData_dict[row[0]] = tuple(row[1:])
        # for key in self.LPTableData_dict:
        #     value2 = float(self.LPTableData_dict[key][0])
        #     value1 = float(self.LPTableData_dict[key][1])
        #     MomInterCurve = np.concatenate((self.PeX.reshape(-1, 1), self.PeY.reshape(-1, 1)), axis=1)
        #     SCF2D = GetSectionCapacityFactor2D.GetSCF2D(MomInterCurve, np.array([value2, value1]))
        #     if SCF2D > 1.0:
        #         tcolor = 'red'
        #     else:
        #         tcolor = 'green'
        #     self.add_point(value2, value1, tcolor)
        #     item = QTableWidgetItem(str(format(SCF2D, '.2f')))
        #     item.setTextAlignment(Qt.AlignCenter)
        #     self.LoadingP_2D_tableWidget.setItem(int(key), 3, item)

        # if column in [2, 3]:
        #     val1 = self.LoadingP_2D_tableWidget.item(row, 2).text()
        #     val2 = self.LoadingP_2D_tableWidget.item(row, 3).text()
        #     print(f"Row {row}, Column 3: {val1}; Column 4: {val2}")
        # def on_cell_changed(self, row, col):
        #     # Check if cell is in row 2, column 2
        #     if row == 1 and col == 1:
        #         item = self.LoadingP_2D_tableWidget.item(row, col)
        #         new_value = item.text()
        #
        #         # Check if new value is a valid number with one decimal place
        #         try:
        #             new_value = float(new_value)
        #             if round(new_value, 1) == new_value:
        #                 print("New value:", new_value)
        #             else:
        #                 item.setText("")
        #         except ValueError:
        #             item.setText("")
        # @Slot(QTableWidgetItem)
        # def LP2DTable_item_changed(self, item):
        # row = item.row()
        # col = item.column()

        # Only calculate if the column is the second column
        # if col == 1:
        #     # Get the values from the first column
        #     val1 = self.LoadingP_2D_tableWidget.item(row, 0).text()
        #     val2 = item.text()
        #
        #     # Perform the calculation
        #     result = float(val1) * float(val2)
        #
        #     # Update the value in the third column
        #     item = QTableWidgetItem(str(result))
        #     item.setFlags(Qt.ItemIsEnabled)
        #     self.LoadingP_2D_tableWidget.setItem(row, 2, item)
        ## Add a new row
        # row_count = self.LoadingP_2D_tableWidget.rowCount()
        # # self.LoadingP_2D_tableWidget.insertRow(row_count)
        # for i in range(row_count):
        #     item = QTableWidgetItem()
        #     item.setText(str(i + 1))
        #     self.LoadingP_2D_tableWidget.setItem(i, 0, item)

    def add_point(self, row, x, y, tcolor):
        #self.LP_scatter = self.ax.scatter(x, y, color='blue', marker='o')
        scatter_points = self.ax.collections
        if row in self.LPTableData_dict.keys():
            tx1 = float(self.LPTableData_dict[row][0])
            ty1 = float(self.LPTableData_dict[row][1])
            for scatter in scatter_points:
                x_data, y_data = scatter.get_offsets().T
                index = np.where((x_data == tx1) & (y_data == ty1))[0]
                if len(index) > 0:
                    scatter.remove()
            # for (tx1, ty1) in scatter_points:
            #     (tx1, ty1).remove()

        self.LPTableData_dict[row] = tuple([x, y])
        self.LP_scatter = self.ax.scatter(x, y, color=tcolor, marker='o')
        ##
        self.ax.autoscale(enable=True, axis='both', tight=False)
        self.ax.relim()
        self.ax.autoscale_view()
        ##
        self.adjust_axes()
        ##
        #self.ax.autoscale(enable=True, axis='both', tight=False)
        # self.ax.relim()
        # self.ax.autoscale_view()
        ##
        self.canvas.draw()
        ##
    def adjust_axes(self):
        ##
        # Get the coordinates of the point created by the plot function
        x_line_coords, y_line_coords = self.Main_curve.get_data()
        # Get the coordinates of the point created by the scatter function
        scatter_points = self.ax.collections
        coords = np.vstack([point.get_offsets() for point in scatter_points])
        x_scatter_coords, y_scatter_coords = coords.T
        # Merge x-axis and y-axis coordinates
        x_coords = np.concatenate((x_line_coords, x_scatter_coords))
        y_coords = np.concatenate((y_line_coords, y_scatter_coords))
        #
        ## Get the maximum value of x-axis and y-axis
        max_x = max(x_coords)
        max_y = max(y_coords)
        ## Get the minimum value of x-axis and y-axis
        min_x = min(x_coords)
        min_y = min(y_coords)
        ## setting the current
        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()
        # # Adjust the axis range according to the new point
        # new_xlim = (min(xlim[0], max_x)-max_x*0.1, max(xlim[1], max_x)+max_x*0.1)
        # new_ylim = (min(ylim[0], max_y)-max_y*0.1, max(ylim[1], max_y)+max_y*0.1)
        new_xlim = (min(min_x, max_x)-max_x*0.1, max(xlim[1], max_x)+max_x*0.1)
        new_ylim = (min(min_y, max_y)-max_y*0.1, max(ylim[1], max_y)+max_y*0.1)
        # Update axis range
        self.ax.set_xlim(new_xlim)
        self.ax.set_ylim(new_ylim)
        ##
    def add_line(self, x, y):
        # x = np.array([1, 2, 3, 4, 5])
        # y = np.array([2, 4, 6, 8, 10])
        self.LP_line, = self.ax.plot(x, y, color='white', linestyle='-', linewidth=0.8, label="LP_line")
        self.lines.append(self.LP_line)
        self.canvas.draw()

    def updat_table_status(self):
        row_count = self.LoadingP_2D_tableWidget.rowCount()
        column_count = self.LoadingP_2D_tableWidget.columnCount()
        for i in range(row_count):
            item = QTableWidgetItem()
            item.setText(str(i + 1))
            # self.LoadingP_2D_tableWidget.setItem(i+1, 0, item)
            self.LoadingP_2D_tableWidget.setItem(i, 0, item)
        ##
        # for j in range(column_count):
        #     item = self.LoadingP_2D_tableWidget.item(0, j)
        #     item.setFlags(item.flags() ^ Qt.ItemIsEditable ^ Qt.ItemIsSelectable)
        for ii in range(row_count):
            item = self.LoadingP_2D_tableWidget.item(ii, 0)
            item.setFlags(item.flags() ^ Qt.ItemIsEditable ^ Qt.ItemIsSelectable)
        ##
        # Center contents of All cell
        for row in range(row_count):
            for column in range(row_count):
                item = self.LoadingP_2D_tableWidget.item(row, column)
                if item is not None:
                    item.setTextAlignment(Qt.AlignCenter)
        ## Save the Loading points data from GUI
        # for i in range(1, self.LoadingP_2D_tableWidget.rowCount()):
        #     row = []
        #     # for j in range(self.LoadingP_2D_tableWidget.columnCount()):
        #     for j in [0, 1, 2]:  ## Loading points
        #         item = self.LoadingP_2D_tableWidget.item(i, j)
        #         if item is not None:
        #             row.append(item.text())
        #         else:
        #             row.append("")
        #     self.LPTableData_dict[row[0]] = tuple(row[1:])
        # for key in self.LPTableData_dict:
        #     value2 = float(self.LPTableData_dict[key][0])
        #     value1 = float(self.LPTableData_dict[key][1])
        #     MomInterCurve = np.concatenate((self.PeX.reshape(-1, 1), self.PeY.reshape(-1, 1)), axis=1)
        #     SCF2D = GetSectionCapacityFactor2D.GetSCF2D(MomInterCurve, np.array([value2, value1]))
        #     if SCF2D > 1.0:
        #         tcolor = 'red'
        #     else:
        #         tcolor = 'green'
        #     self.add_point(value2, value1, tcolor)
        #     item = QTableWidgetItem(str(format(SCF2D, '.2f')))
        #     item.setTextAlignment(Qt.AlignCenter)
        #     self.LoadingP_2D_tableWidget.setItem(int(key), 3, item)
            # self.add_point(self, x, y, tcolor):


    def updat_figure_status(self):
        return

    def on_cell_clicked(self, row, column):
        # item1 = self.LoadingP_2D_tableWidget.item(row, 1)
        # item2 = self.LoadingP_2D_tableWidget.item(row, 2)
        if row == self.LoadingP_2D_tableWidget.rowCount() - 1:
        #if row == self.LoadingP_2D_tableWidget.rowCount() - 1 and item1.text()!='' and item2.text()!='':
            self.LoadingP_2D_tableWidget.insertRow(self.LoadingP_2D_tableWidget.rowCount())
        ##
        # item = self.LoadingP_2D_tableWidget.item(row, column)
        # editor = QLineEdit(self)
        # validator = QDoubleValidator(self)
        # editor.setValidator(validator)
        # # editor.editingFinished.connect(lambda: self.finish_editing(editor, item))
        # self.LoadingP_2D_tableWidget.setCellWidget(row, column, editor)
        # editor.setFocus()
        # editor.selectAll()
        # editor.setStyleSheet("border: none;")  # remove border
        ##
        self.updat_table_status()
    def YS2DPlot(self):
        if self.grandparent_window.Centerline_radioButton.isChecked():
            YieldSAnalResults = CMYieldSAnalResults
        elif self.grandparent_window.Outline_radioButton.isChecked():
            YieldSAnalResults = FEYieldSAnalResults
        if self.parent_window.PYSPMy_radioButton.isChecked():
            ploteY = YieldSAnalResults.ONx_y
            ploteX = YieldSAnalResults.OMy_x
            # tploteY = ploteY.values()
            # tploteX = ploteX.values()
            # ##
            # PeY = np.array(list(tploteY))
            # PeX = np.array(list(tploteX))
            ##
        elif self.parent_window.PYSPMz_radioButton.isChecked():
            ploteY = YieldSAnalResults.ONx_z
            ploteX = YieldSAnalResults.OMz_x
            # tploteY = ploteY.values()
            # tploteX = ploteX.values()
            # ##
            # PeY = np.array(list(tploteY))
            # PeX = np.array(list(tploteX))
        elif self.parent_window.PYSMyMz_radioButton.isChecked():
            ploteY = YieldSAnalResults.OMy_z
            ploteX = YieldSAnalResults.OMz_y
            # tploteY = YieldSAnalResults.OMy_z
            # tploteX = YieldSAnalResults.OMz_y
            # ploteY = {}
            # ploteX = {}
            # tMstep = YieldSurfaceAnalInfo.MStep
            # for i, (k, v) in enumerate(tploteY.items()):
            #     ploteY[k] = v
            #     if i == tMstep: break
            # for i, (k, v) in enumerate(tploteX.items()):
            #     ploteX[k] = v
            #     if i == tMstep: break
            # # tploteY = ploteY.values()
            # # tploteX = ploteX.values()
            # ##
            # print("Nx_yz = ", np.array(list(YieldSAnalResults.ONx_yz.values())))
            # tPeY = np.array(list(ploteY.values()))
            # tPeX = np.array(list(ploteX.values()))
            ##
        tploteY = ploteY.values()
        tploteX = ploteX.values()
        ##
        self.PeY = np.array(list(tploteY))
        self.PeX = np.array(list(tploteX))
        ##for testing
        #MomInterCurve = np.concatenate((self.PeX.reshape(-1, 1), self.PeY.reshape(-1, 1)), axis=1)
        # elif self.mw.PYSMyMz_radioButton.isChecked():
        #
        font1 = {'family': 'Times New Roman', 'size': 10, 'weight': 'bold'}
        font2 = {'family': 'Times New Roman', 'size': 10, 'weight': 'bold'}
        #
        if self.parent_window.PYSPMy_radioButton.isChecked() and self.parent_window.PrincipalAxis_radioButton.isChecked():
            self.ax.set_xlabel('Bending Moment Mv', color='white', fontsize=10, fontdict=font1)
            self.ax.set_ylabel('Axial Force Px', color='white', fontsize=10, fontdict=font1)
            self.TableHX = "Mv"
            self.TableHY = "Px"
            self.SheetName = "Px vs. Mv"
            self.ax.set_title('Px vs. Mv', color='white', fontsize=10, fontdict=font2)
            self.PeY = np.array(list(tploteY))
        elif self.parent_window.PYSPMy_radioButton.isChecked() and self.parent_window.UsedefinedAxis_radioButton.isChecked():
            self.ax.set_xlabel('Bending Moment My', color='white', fontsize=10, fontdict=font1)
            self.ax.set_ylabel('Axial Force Px', color='white', fontsize=10, fontdict=font1)
            self.TableHX = "My"
            self.TableHY = "Px"
            self.SheetName = "Px vs. My"
            self.ax.set_title('Px vs. My', color='white', fontsize=10, fontdict=font2)
            self.PeY = np.array(list(tploteY))
        elif self.parent_window.PYSPMz_radioButton.isChecked() and self.parent_window.PrincipalAxis_radioButton.isChecked():
            self.ax.set_xlabel('Bending Moment Mw', color='white', fontsize=10, fontdict=font1)
            self.ax.set_ylabel('Axial Force Px', color='white', fontsize=10, fontdict=font1)
            self.TableHX = "Mw"
            self.TableHY = "Px"
            self.SheetName = "Px vs. Mw"
            self.ax.set_title('Px vs. Mw', color='white', fontsize=10, fontdict=font2)
            self.PeY = np.array(list(tploteY))
        elif self.parent_window.PYSPMz_radioButton.isChecked() and self.parent_window.UsedefinedAxis_radioButton.isChecked():
            self.ax.set_xlabel('Bending Moment Mz', color='white', fontsize=10, fontdict=font1)
            self.ax.set_ylabel('Axial Force Px', color='white', fontsize=10, fontdict=font1)
            self.TableHX = "Mz"
            self.TableHY = "Px"
            self.SheetName = "Px vs. Mz"
            self.ax.set_title('Px vs. Mz', color='white', fontsize=10, fontdict=font2)
            self.PeY = np.array(list(tploteY))
        elif self.parent_window.PYSMyMz_radioButton.isChecked() and self.parent_window.PrincipalAxis_radioButton.isChecked():
            self.ax.set_xlabel('Bending Moment Mw', color='white', fontsize=10, fontdict=font1)
            self.ax.set_ylabel('Bending Moment Mv', color='white', fontsize=10, fontdict=font1)
            self.ax.set_title('Moment Interaction Curve\n (Px = 0.0)', color='white', fontsize=9)
            self.TableHX = "Mw"
            self.TableHY = "Mv"
            self.SheetName = "Mv vs. Mw"
            self.ax.set_title('Mv vs. Mw', color='white', fontsize=10, fontdict=font2)
        elif self.parent_window.PYSMyMz_radioButton.isChecked() and self.parent_window.UsedefinedAxis_radioButton.isChecked():
            self.ax.set_xlabel('Bending Moment Mz', color='white', fontsize=10, fontdict=font1)
            self.ax.set_ylabel('Bending Moment My', color='white', fontsize=10, fontdict=font1)
            self.ax.set_title('Moment Interaction Curve\n (Px = 0.0)', color='white', fontsize=9)
            self.TableHX = "Mz"
            self.TableHY = "My"
            self.SheetName = "My vs. Mz"
            self.ax.set_title('My vs. Mz', fontsize=10, color='white', fontdict=font2)
        ##
        # Create the subplot
        #ax = self.fig2D.add_subplot(111)
        # Plot the data
        self.Main_curve, = self.ax.plot(self.PeX, self.PeY, label="Main_curve", color="red")
        ## For testing
        # df1 = pd.DataFrame.from_dict(list(tploteY))
        # df2 = pd.DataFrame.from_dict(list(tploteX))
        #
        # with pd.ExcelWriter('C:/Users/gaowl/PycharmProjects/Mastan3/Source/gui/msasect/examples/output_OTTTMy_z.xlsx') as writer:
        #     # 将 DataFrame 写入 Excel 文件
        #     df1.to_excel(writer, sheet_name='Sheet1', index=False)
        #
        # with pd.ExcelWriter('C:/Users/gaowl/PycharmProjects/Mastan3/Source/gui/msasect/examples/output_OTTTMz_y.xlsx') as writer:
        #     # 将 DataFrame 写入 Excel 文件
        #     df2.to_excel(writer, sheet_name='Sheet1', index=False)

        self.lines.append(self.Main_curve)
        self.ax.spines['bottom'].set_linewidth(1.8)
        self.ax.spines['left'].set_linewidth(1.8)
        ##
        #self.ax.set_facecolor('#FFFFE0')
        self.ax.set_facecolor("black")
        self.ax.tick_params(axis="x", colors="white", labelsize=10)
        self.ax.tick_params(axis="y", colors="white", labelsize=10)
        self.ax.spines['bottom'].set_color('white')
        self.ax.spines['top'].set_color('white')
        self.ax.spines['right'].set_color('white')
        self.ax.spines['left'].set_color('white')
        # for label in self.ax.get_xticklabels() + self.ax.get_yticklabels():
        #     label.set_fontname('Times New Roman')
        # self.ax.set_facecolor((43/255, 43/255, 43/255))
        # Set the labels and title
        self.ax.tick_params(direction='in')
        #plt.rcParams['axes.labelsize'] = 'large'
        plt.tick_params(axis='both', which='major', labelsize=6, pad=6)
        plt.tight_layout()
        ##
        # self.ax.set_title('Moment Interaction Curve', fontsize=9)
        self.ax.grid(linestyle=':', linewidth='0.8', color='gray', alpha=0.8)
        ##
        # # get font properties and set size
        # font = font_manager.FontProperties()
        # font.set_size(self.fig2D.dpi * 0.1)
        #
        # # set font properties for x and y-axis labels
        # self.ax.xaxis.label.set_font_properties(font)
        # self.ax.yaxis.label.set_font_properties(font)
        # Set the coordinate axis number format to fixed-point notation.
        # self.ax.xaxis.set_major_formatter(plt.FormatStrFormatter('%.1f'))
        # self.ax.yaxis.set_major_formatter(plt.FormatStrFormatter('%.1f'))
        self.ax.xaxis.set_major_formatter(plt.FormatStrFormatter('%.1e'))
        self.ax.yaxis.set_major_formatter(plt.FormatStrFormatter('%.1e'))
        for label in self.ax.get_xticklabels() + self.ax.get_yticklabels():
            label.set_fontname('Times New Roman')
        ##
        self.fig2D.tight_layout(pad=1.08, h_pad=None, w_pad=None, rect=None)

    @Slot()
    def on_ExportSP_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError

        # # Example data
        # x = np.arange(0, 10, 0.1)
        # y = np.sin(x)

        # Create the Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = self.SheetName

        # Write the data to the worksheet
        ws.cell(row=1, column=1, value=self.TableHX)
        ws.cell(row=1, column=2, value=self.TableHY)
        for i in range(len(self.PeX)):
            ws.cell(row=i + 2, column=1, value=self.PeX[i])
            ws.cell(row=i + 2, column=2, value=self.PeY[i])

        # Open a dialog to choose the save location
        save_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save Excel File", str("Sect_"+ws.title), "Excel Files (*.xlsx)")

        if save_path:
            # Save the workbook to the chosen location
            wb.save(save_path)

    @Slot()
    def on_ClearInput_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        #raise NotImplementedError
        #self.LoadingP_2D_tableWidget.clearContents()
        scatter_points = self.ax.collections
        for scatter in scatter_points:
            scatter.remove()
        ##
        self.ax.autoscale_view()
        self.ax.autoscale(enable=True, axis='both', tight=False)
        self.ax.relim()
        #
        self.canvas.draw()
        ##
        self.initLoadingPTable()
        #
        for row in range(1, self.LoadingP_2D_tableWidget.rowCount()):
            for col in range(1, self.LoadingP_2D_tableWidget.columnCount()):
                item = self.LoadingP_2D_tableWidget.item(row, col)
                if item is not None:
                    self.LoadingP_2D_tableWidget.setItem(row, col, None)
        ##
        self.updat_table_status()

        ##
        # num_lines = len(self.lines)
        # print("Number of lines:", num_lines)
        # for ii in range(1, num_lines):
        #     try:
        #         Remove_line = self.lines[ii]
        #         Remove_line.remove()
        #         self.lines.remove(Remove_line)
        #     except IndexError:
        #         print("Index out of range")

        # self.ax.lines.remove(self.LP_line)
        # self.LP_line.remove()



    @Slot()
    def on_Close_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        ## raise NotImplementedError
        QDialog.close(self)

    @Slot()
    def on_Delete_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError
        # selected_rows = self.LoadingP_2D_tableWidget.selectedItems()
        # if selected_rows:  #
        currentRow = self.LoadingP_2D_tableWidget.currentRow()
        #
        scatter_points = self.ax.collections
        if currentRow != -1 and currentRow != 0:
            self.LoadingP_2D_tableWidget.removeRow(currentRow)
            self.LoadingP_2D_tableWidget.clearSelection()
            ##
            if currentRow in self.LPTableData_dict.keys():
                tx1 = float(self.LPTableData_dict[currentRow][0])
                ty1 = float(self.LPTableData_dict[currentRow][1])
                for scatter in scatter_points:
                    x_data, y_data = scatter.get_offsets().T
                    index = np.where((x_data == tx1) & (y_data == ty1))[0]
                    if len(index) > 0:
                        scatter.remove()
                    #
                    self.canvas.draw()
                    ##
                    self.ax.autoscale(enable=True, axis='both', tight=False)
                    self.ax.relim()
                    self.ax.autoscale_view()
                    ##
                    self.adjust_axes()
        ##
        self.updat_table_status()
