from PySide6.QtCore import Slot
from PySide6.QtGui import QIcon, QFont
from openpyxl import Workbook
import numpy as np
from gui.msasect.ui.Ui_ShowResultsBuckling import Ui_GlobalBucklingPlot_Dialog
from PySide6.QtWidgets import  QDialog,  QFileDialog
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from gui.msasect.base.Model import GlobalBuckling
import matplotlib.pyplot as plt
from analysis.GlobalBuckling.variables import Model as GlobalBucklingmodel
import pyqtgraph as pg
from pyqtgraph import AxisItem
from pyqtgraph.Qt import QtGui, QtCore

class GlobalBucklingPlot_Dialog(QDialog, Ui_GlobalBucklingPlot_Dialog):
    """
    Class documentation goes here.
    """
    def __init__(self, mw, parent=None):
        """
        Constructor

        @param parent reference to the parent widget (defaults to None)
        @type QWidget (optional)
        """
        super().__init__(parent)
        self.setupUi(self)
        self.mw = mw
        self.setWindowIcon(QIcon('ui/ico/GlobalBuckling.png'))
        self.data = {}
        self.PlotBucklingCurves_Compression()
        self.PlotBucklingCurves_MomentPositive()
        self.PlotBucklingCurves_MomentNegative()
        self.PeX = np.array([])


    def PlotBucklingCurves_Compression (self):
        for i in reversed(range(self.Compression_Layout.count())):
            self.Compression_Layout.itemAt(i).widget().setParent(None)
        Buckling_data = GlobalBuckling.Buckling_data
        if Buckling_data['method'] == 'Analytical':
            A = GlobalBucklingmodel.SecProperties.Area
            Iv = GlobalBucklingmodel.SecProperties.MomentofInertia_v
            Iw = GlobalBucklingmodel.SecProperties.MomentofInertia_w
            I = min(Iv, Iw)
            vs = GlobalBucklingmodel.SecProperties.ShearCentre_v
            ws = GlobalBucklingmodel.SecProperties.ShearCentre_w
            r = np.abs(np.sqrt((Iv + Iw) / A + vs**2 + ws**2))
            Lamuda = Buckling_data['Lamuda_method4']
            Pfb1 = []
            Pfb2 = []
            Pfb3 = []
            if len(Buckling_data ['Pfb1'])>0 :
                Pfb1 = np.array(Buckling_data ['Pfb1'])
            if len(Buckling_data ['Pfb2'])>0:
                Pfb2 = np.array(Buckling_data ['Pfb2'])
            if len(Buckling_data ['Pfb3'])>0:
                Pfb3 = np.array(Buckling_data ['Pfb3'])
            Pcr1 = self.solve_cubic(Pfb1, Pfb2, Pfb3[0], r, vs, ws)
            if np.iscomplex(Pcr1).any():
                Pcr1 = self.find_min_elements(Pfb1, Pfb2, Pfb3[0],[])
            # for i in range(len(Pcr1)):
            #     Pcr1[i] = Pcr1[i][2]
            if Buckling_data['Material'] == 'Elastic' or Buckling_data['Material'] == 'Plastic':
                Pcr1 = self.find_min_elements(Pcr1, Buckling_data['Elastic compression'], [], [])
            elif Buckling_data['Material'] == 'Ignore':
                Pcr1 = self.find_min_elements(Pcr1, [], [], [])
            # if Buckling_data['Material'] == 'Elastic':
            #     Pcr1 = self.find_min_elements(Pcr1, Buckling_data['Elastic bending'], [], [])
            # elif Buckling_data['Material'] == 'Plastic':
            #     Pcr1 = self.find_min_elements(Pcr1, Buckling_data['Plastic bending'], [], [])
            Pcr2 = self.solve_cubic(Pfb1, Pfb2, Pfb3[1], r, vs, ws)
            if np.iscomplex(Pcr2).any():
                Pcr2 = self.find_min_elements(Pfb1, Pfb2, Pfb3[1],[])
            # for i in range(len(Pcr2)):
            #     Pcr2[i] = Pcr2[i][2]
            # if Buckling_data['Material'] == 'Elastic':
            #     Pcr2 = self.find_min_elements(Pcr2, Buckling_data['Elastic bending'], [], [])
            # elif Buckling_data['Material'] == 'Plastic':
            #     Pcr2 = self.find_min_elements(Pcr2, Buckling_data['Plastic bending'], [], [])
            if Buckling_data['Material'] == 'Elastic' or Buckling_data['Material'] == 'Plastic':
                Pcr2 = self.find_min_elements(Pcr2, Buckling_data['Elastic compression'], [], [])
            elif Buckling_data['Material'] == 'Ignore':
                Pcr2 = self.find_min_elements(Pcr2, [], [], [])
            self.data['Pcr1'] = Pcr1
            self.data['Pcr2'] = Pcr2
            # print(Buckling_data['Material'])
            self.plot = pg.PlotWidget()
            self.Compression_Layout.addWidget(self.plot)
            self.plot.setBackground('black')
            self.plot.showGrid(x=True, y=True)
            self.curve1 = self.plot.plot(Lamuda, Pcr1, pen=pg.mkPen(color='red', width=3) )
            self.curve2 = self.plot.plot(Lamuda, Pcr2, pen=pg.mkPen(color='yellow', width=3,style=QtCore.Qt.DashLine))
            self.curvea = self.plot.plot([0, 0.0001], [max(Pcr1)*0.90, max(Pcr1)*1.1], pen=pg.mkPen(color='w', width=0.01) )
            Label_style = {'font-size': '12pt', 'font-family': 'Times New Roman'}
            font = QtGui.QFont()
            font.setFamily('Times')
            font.setPointSize(12)
            self.plot.setTitle("Global Buckling Curves - Compression", color='w', font = font)
            self.plot.setLabel("left", "Compression", **Label_style)
            self.plot.setLabel("bottom", "λ", **Label_style)
            self.plot.showAxis('top')
            self.plot.showAxis('right')
            self.plot.getAxis('top').setTicks([])
            self.plot.getAxis('right').setTicks([])
            self.line = pg.InfiniteLine(angle=90, movable=True, pen='gray', label='')
            self.plot.addItem(self.line)
            self.annotation = pg.TextItem(text='', anchor=(0, 1))
            self.plot.addItem(self.annotation)

            def on_mouse_move(evt):
                if self.plot.sceneBoundingRect().contains(evt):
                    mouse_point = self.plot.getViewBox().mapSceneToView(evt)
                    x = mouse_point.x()
                    if x >= Lamuda[0] and x <= Lamuda[-1]:
                        # index = np.searchsorted(Lamuda, x)
                        P1 = float(np.interp(x, Lamuda, Pcr1))
                        P2 = float(np.interp(x, Lamuda, Pcr2))
                        y_diff = (P2 - P1) / P2
                        self.line.setValue(x)
                        self.annotation.setText('L = {:.2f}\nP1 = {:.3e}\nP2 = {:.3e}\nDiff = (P1-P2)/P2 ={:.2f}%'.format(x*np.sqrt(I/A),P1,P2,y_diff*100), color=(128, 128, 128))
                        self.annotation.setPos(x, 0.92*np.mean([P1, P2]))

            self.plot.scene().sigMouseMoved.connect(on_mouse_move)
            # Set the background color of the PlotWidget to black
            self.plot.setStyleSheet("background-color: #000000;")

            # Set the color of the axis and labels to white
            styles = {'color': 'w'}
            # set x_axis
            x_axis = self.plot.getAxis('bottom')
            x_axis.setTextPen(pg.mkPen(**styles))
            x_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
            x_axis.setPen(pg.mkPen(width=2))
            # set y_axis
            y_axis = self.plot.getAxis('left')
            y_axis.enableAutoSIPrefix(True)
            y_axis.setTextPen(pg.mkPen(**styles))
            y_axis.setTickFont(QtGui.QFont("Times New Roman", 12))
            y_axis.setPen(pg.mkPen(width=2))
            # other axis
            axis = self.plot.getAxis('right')
            axis.setPen(pg.mkPen(**styles))
            axis.setPen(pg.mkPen(width=2))
            axis = self.plot.getAxis('top')
            axis.setPen(pg.mkPen(**styles))
            axis.setPen(pg.mkPen(width=2))
            #Legend
            legend = pg.LegendItem()
            legend.setGeometry(QtCore.QRectF(0, 0, 150, 100))
            legend.setParentItem(self.plot.graphicsItem())
            legend.addItem(self.curve1, 'Considering twisting effects (P1)')
            legend.addItem(self.curve2, 'Ignoring twisting effects (P2)')
            legend.setAutoFillBackground(True)
            legend.setBrush('k')
            legend.anchor((1, 0), (1, 0), offset=(-30, 40))



    def PlotBucklingCurves_MomentPositive(self):
        for i in reversed(range(self.NegativeBending_Layout.count())):
            self.NegativeBending_Layout.itemAt(i).widget().setParent(None)
        A = GlobalBucklingmodel.SecProperties.Area
        Iv = GlobalBucklingmodel.SecProperties.MomentofInertia_v
        Iw = GlobalBucklingmodel.SecProperties.MomentofInertia_w
        I = min(Iv, Iw)
        Buckling_data = GlobalBuckling.Buckling_data
        Lamuda2 = Buckling_data['Lamuda_method4']
        if len(Buckling_data['Pfb5']) != 0:
            Pfb5 = Buckling_data['Pfb5']
            if Buckling_data['Material'] == 'Elastic':
                Glb5_P = self.find_min_elements(Buckling_data['Pfb5'], Buckling_data['Elastic bending'], [], [])
            elif Buckling_data['Material'] == 'Plastic':
                Glb5_P = self.find_min_elements(Buckling_data['Pfb5'], Buckling_data['Plastic bending'], [], [])
            elif Buckling_data['Material'] == 'Ignore':
                Glb5_P = Buckling_data['Pfb5']
        if len(Buckling_data ['Pfb4'])!=0:
            Pfb4 = Buckling_data ['Pfb4']
            if Buckling_data['Material'] == 'Elastic':
                PositiveBending = self.find_min_elements(Pfb4[0], Buckling_data['Elastic bending'], [], [])
            elif Buckling_data['Material'] == 'Plastic':
                PositiveBending = self.find_min_elements(Pfb4[0], Buckling_data['Plastic bending'], [], [])
            elif Buckling_data['Material'] == 'Ignore':
                PositiveBending = self.find_min_elements(Pfb4[0], [], [], [])
        self.plot2 = pg.PlotWidget()
        self.NegativeBending_Layout.addWidget(self.plot2)
        self.plot2.setBackground('black')
        self.plot2.showGrid(x=True, y=True)
        self.curve3 = self.plot2.plot(Lamuda2, PositiveBending, pen=pg.mkPen(color='red', width=3), )
        self.curve4 = self.plot2.plot(Lamuda2, np.array(Glb5_P), pen=pg.mkPen(color='yellow', width=3, style=QtCore.Qt.DashLine))
        self.curveb = self.plot2.plot([0, 0.0001], [max(Glb5_P)*0.90, max(Glb5_P)*1.1], pen=pg.mkPen(color='w', width=0.01))
        #print(Pfb4)
        self.data['Glb5_P'] = Glb5_P
        self.data['NegativeBending'] = PositiveBending
        Label_style = {'font-size': '12pt', 'font-family': 'Times New Roman'}
        font = QtGui.QFont()
        font.setFamily('Times')
        font.setPointSize(12)
        self.plot2.setTitle("Global Buckling Curves - Negative Bending", color='w', font=font)
        self.plot2.setLabel("left", "Negative Bending", **Label_style)
        self.plot2.setLabel("bottom", "λ", **Label_style)
        self.plot2.showAxis('top')
        self.plot2.showAxis('right')
        self.plot2.getAxis('top').setTicks([])
        self.plot2.getAxis('right').setTicks([])
        self.line2 = pg.InfiniteLine(angle=90, movable=True, pen='gray', label='')
        self.plot2.addItem(self.line2)
        self.annotation2 = pg.TextItem(text='', anchor=(0, 1))
        self.plot2.addItem(self.annotation2)

        def on_mouse_move2(evt):
            if self.plot2.sceneBoundingRect().contains(evt):
                mouse_point2 = self.plot2.getViewBox().mapSceneToView(evt)
                x2 = mouse_point2.x()
                if x2 >= Lamuda2[0] and x2 <= Lamuda2[-1]:
                    # index = np.searchsorted(Lamuda2, x2)
                    Positive = float(np.interp(x2, Lamuda2, PositiveBending))
                    Glb = float(np.interp(x2, Lamuda2, np.array(Glb5_P)))
                    y_diff2 = (Positive - Glb) / Glb
                    self.line2.setValue(x2)
                    self.annotation2.setText('L = {:.2f}\nM1 = {:.3e}\nM2 = {:.3e}\nDiff = (M1-M2)/M2 ={:.2f}%'.format(x2*np.sqrt(I/A),Positive,Glb,y_diff2 * 100), color=(128, 128, 128))
                    self.annotation2.setPos(x2, 0.92*np.mean([Positive, Glb]))

        self.plot2.scene().sigMouseMoved.connect(on_mouse_move2)
        # Set the background color of the PlotWidget to black
        self.plot2.setStyleSheet("background-color: #000000;")

        # Set the color of the axis and labels to white
        styles = {'color': 'w'}
        # set x_axis
        x_axis2 = self.plot2.getAxis('bottom')
        x_axis2.setTextPen(pg.mkPen(**styles))
        x_axis2.setTickFont(QtGui.QFont("Times New Roman", 12))
        x_axis2.setPen(pg.mkPen(width=2))
        # set y_axis
        y_axis2 = self.plot2.getAxis('left')
        y_axis2.enableAutoSIPrefix(True)
        y_axis2.setTextPen(pg.mkPen(**styles))
        y_axis2.setTickFont(QtGui.QFont("Times New Roman", 12))
        y_axis2.setPen(pg.mkPen(width=2))
        # other axis
        axis2 = self.plot2.getAxis('right')
        axis2.setPen(pg.mkPen(**styles))
        axis2.setPen(pg.mkPen(width=2))
        axis2 = self.plot2.getAxis('top')
        axis2.setPen(pg.mkPen(**styles))
        axis2.setPen(pg.mkPen(width=2))
        # Legend
        legend2 = pg.LegendItem()
        legend2.setGeometry(QtCore.QRectF(0, 0, 150, 100))
        legend2.setParentItem(self.plot2.graphicsItem())
        legend2.addItem(self.curve3, 'Considering twisting effects (M1)')
        legend2.addItem(self.curve4, 'Ignoring twisting effects (M2)')
        legend2.setAutoFillBackground(True)
        legend2.setBrush('k')
        legend2.anchor((1, 0), (1, 0), offset=(-30, 40))



    def PlotBucklingCurves_MomentNegative(self):
        for i in reversed(range(self.PositiveBending_Layout.count())):
            self.PositiveBending_Layout.itemAt(i).widget().setParent(None)
        A = GlobalBucklingmodel.SecProperties.Area
        Iv = GlobalBucklingmodel.SecProperties.MomentofInertia_v
        Iw = GlobalBucklingmodel.SecProperties.MomentofInertia_w
        I = min(Iv, Iw)
        Buckling_data = GlobalBuckling.Buckling_data
        Lamuda3 = Buckling_data['Lamuda_method4']
        if len(Buckling_data['Pfb4']) != 0:
            Pfb4 = Buckling_data['Pfb4']
            if len(Buckling_data['Pfb5']) != 0:
                Glb5_N = Buckling_data['Pfb5']
                if Buckling_data['Material'] == 'Elastic':
                    Glb5_N = self.find_min_elements(Buckling_data['Pfb5'], Buckling_data['Elastic bending'], [], [])
                elif Buckling_data['Material'] == 'Plastic':
                    Glb5_N = self.find_min_elements(Buckling_data['Pfb5'], Buckling_data['Plastic bending'], [], [])
                elif Buckling_data['Material'] == 'Ignore':
                    Glb5_N = self.find_min_elements(Buckling_data['Pfb5'], [], [], [])
            if Buckling_data['Material'] == 'Elastic':
                NegativeBending = self.find_min_elements(Pfb4[1], Buckling_data['Elastic bending'], [], [])
            elif Buckling_data['Material'] == 'Plastic':
                NegativeBending = self.find_min_elements(Pfb4[1], Buckling_data['Plastic bending'], [], [])
            elif Buckling_data['Material'] == 'Ignore':
                NegativeBending = self.find_min_elements(Pfb4[1], [], [], [])
        self.plot3 = pg.PlotWidget()
        self.PositiveBending_Layout.addWidget(self.plot3)
        self.plot3.setBackground('black')
        self.plot3.showGrid(x=True, y=True)
        self.curve5 = self.plot3.plot(Lamuda3, NegativeBending, pen=pg.mkPen(color='red', width=3), )
        self.curve6 = self.plot3.plot(Lamuda3, np.array(Glb5_N),pen=pg.mkPen(color='yellow', width=3, style=QtCore.Qt.DashLine))
        self.curvec = self.plot3.plot([0,0.0001], [max(Glb5_N)*0.9, max(Glb5_N)*1.1], pen=pg.mkPen(color='w', width=0.01))
        # print(Pfb4)
        self.data['Glb5_N'] = Glb5_N
        self.data['PositiveBending'] = NegativeBending
        Label_style = {'font-size': '12pt', 'font-family': 'Times New Roman'}
        font = QtGui.QFont()
        font.setFamily('Times')
        font.setPointSize(12)
        self.plot3.setTitle("Global Buckling Curves - Positive Bending", color='w', font=font)
        self.plot3.setLabel("left", "Positive Bending", **Label_style)
        self.plot3.setLabel("bottom", "λ", **Label_style)
        self.plot3.showAxis('top')
        self.plot3.showAxis('right')
        self.plot3.getAxis('top').setTicks([])
        self.plot3.getAxis('right').setTicks([])
        self.line3 = pg.InfiniteLine(angle=90, movable=True, pen='gray', label='')
        self.plot3.addItem(self.line3)
        self.annotation3 = pg.TextItem(text='', anchor=(0, 1))
        self.plot3.addItem(self.annotation3)

        def on_mouse_move3(evt):
            if self.plot3.sceneBoundingRect().contains(evt):
                mouse_point3 = self.plot3.getViewBox().mapSceneToView(evt)
                x3 = mouse_point3.x()
                if x3 >= Lamuda3[0] and x3 <= Lamuda3[-1]:
                    # index = np.searchsorted(Lamuda3, x3)
                    Negative = float(np.interp(x3, Lamuda3, NegativeBending))
                    Gb5 = float(np.interp(x3, Lamuda3, np.array(Glb5_N)))
                    y_diff3 = (Negative - Gb5) / Gb5
                    self.line3.setValue(x3)
                    self.annotation3.setText('L = {:.2f}\nM1 = {:.3e}\nM2 = {:.3e}\nDiff = (M1-M2)/M2 ={:.2f}%'.format(x3*np.sqrt(I/A),Negative,Gb5,y_diff3 * 100), color=(128, 128, 128))
                    self.annotation3.setPos(x3, 0.92*np.mean([Negative, Gb5]))

        self.plot3.scene().sigMouseMoved.connect(on_mouse_move3)
        # Set the background color of the PlotWidget to black
        self.plot3.setStyleSheet("background-color: #000000;")

        # Set the color of the axis and labels to white
        styles = {'color': 'w'}
        # set x_axis
        x_axis3 = self.plot3.getAxis('bottom')
        x_axis3.setTextPen(pg.mkPen(**styles))
        x_axis3.setTickFont(QtGui.QFont("Times New Roman", 12))
        x_axis3.setPen(pg.mkPen(width=2))
        # set y_axis
        y_axis3 = self.plot3.getAxis('left')
        y_axis3.enableAutoSIPrefix(True)
        y_axis3.setTextPen(pg.mkPen(**styles))
        y_axis3.setTickFont(QtGui.QFont("Times New Roman", 12))
        y_axis3.setPen(pg.mkPen(width=2))
        # other axis
        axis3 = self.plot3.getAxis('right')
        axis3.setPen(pg.mkPen(**styles))
        axis3.setPen(pg.mkPen(width=2))
        axis3 = self.plot3.getAxis('top')
        axis3.setPen(pg.mkPen(**styles))
        axis3.setPen(pg.mkPen(width=2))
        # Legend
        legend3 = pg.LegendItem()
        legend3.setGeometry(QtCore.QRectF(0, 0, 150, 100))
        legend3.setParentItem(self.plot3.graphicsItem())
        legend3.addItem(self.curve5, 'Considering twisting effects (M1)')
        legend3.addItem(self.curve6, 'Ignoring twisting effects (M2)')
        legend3.setAutoFillBackground(True)
        legend3.setBrush('k')
        legend3.anchor((1, 0), (1, 0), offset=(-30, 40))

    @Slot()
    def on_ExportResults_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError
        Buckling_data = GlobalBuckling.Buckling_data
        if Buckling_data['method'] == 'Analytical':
            wb = Workbook()
            sheet1 = wb.active
            sheet1.title = 'Compression'
            # Write the data to the worksheet
            Lamuda = np.array(Buckling_data['Lamuda_method4'])
            sheet1.cell(row=1, column=1, value='λ')
            for i in range(len(Lamuda)):
                sheet1.cell(row=i + 2, column=1, value=Lamuda[i])
            ii = 2
            if len(self.data['Pcr1']) > 0:
                Pfb1 = np.array(self.data['Pcr1'])
                sheet1.cell(row=1, column=ii, value='Considering Twisting Effects')
                for i in range(len(Pfb1)):
                    sheet1.cell(row=i + 2, column=ii, value=Pfb1[i])
                ii += 1
            if len(self.data['Pcr2']) > 0:
                Pfb2 = np.array(self.data['Pcr2'])
                sheet1.cell(row=1, column=ii, value='Ignoring Twisting Effects')
                for i in range(len(Pfb2)):
                    sheet1.cell(row=i + 2, column=ii, value=Pfb2[i])
                ii += 1

            sheet2 = wb.create_sheet(title='Positive Bending')
            sheet2.cell(row=1, column=1, value='λ')
            for i in range(len(Lamuda)):
                sheet2.cell(row=i + 2, column=1, value=Lamuda[i])
            ii = 2
            if len(self.data['Glb5_N']) > 0:
                Glb5_N = np.array(self.data['Glb5_N'])
                sheet2.cell(row=1, column=ii, value='Ignoring Twisting Effects')
                for i in range(len(Glb5_N)):
                    sheet2.cell(row=i + 2, column=ii, value=Glb5_N[i])
                ii += 1
            if len(self.data['NegativeBending']) > 0:
                GlbN = np.array(self.data['NegativeBending'])
                sheet2.cell(row=1, column=ii, value='Considering Twisting Effects')
                for i in range(len(GlbN)):
                    sheet2.cell(row=i + 2, column=ii, value=GlbN[i])
                ii += 1

            sheet3 = wb.create_sheet(title='Negative Bending')
            sheet3.cell(row=1, column=1, value='λ')
            for i in range(len(Lamuda)):
                sheet3.cell(row=i + 2, column=1, value=Lamuda[i])
            ii = 2
            if len(self.data['Glb5_P']) > 0:
                Glb5_P = np.array(self.data['Glb5_P'])
                sheet3.cell(row=1, column=ii, value='Ignoring Twisting Effects')
                for i in range(len(Glb5_P)):
                    sheet3.cell(row=i + 2, column=ii, value=Glb5_P[i])
                ii += 1
            if len(self.data['PositiveBending']) > 0:
                GlbN = np.array(self.data['PositiveBending'])
                sheet3.cell(row=1, column=ii, value='Considering Twisting Effects')
                for i in range(len(GlbN)):
                    sheet3.cell(row=i + 2, column=ii, value=GlbN[i])
                ii += 1

            # Open a dialog to choose the save location
            save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", str("Global Buckling Sect_"), "Excel Files (*.xlsx)")
        elif Buckling_data['method'] == 'Line_element':
            wb = Workbook()
            sheet1 = wb.active
            sheet1.title = 'Eigen-buckling analysis'
            # Write the data to the worksheet
            Lamuda = np.array(Buckling_data['Lamuda'])
            sheet1.cell(row=1, column=1, value='λ')
            for i in range(len(Lamuda)):
                sheet1.cell(row=i + 2, column=1, value=Lamuda[i])
            ii = 2
            if len(Buckling_data['Factors1']) > 0:
                LF1 = np.array(Buckling_data['Factors1'])
                sheet1.cell(row=1, column=ii, value='Load Factor (Mode NO.1)')
                for i in range(len(LF1)):
                    sheet1.cell(row=i + 2, column=ii, value=LF1[i])
                ii += 1
            # if len(Buckling_data['Factors2']) > 0:
            #     LF2 = np.array(Buckling_data['Factors2'])
            #     sheet1.cell(row=1, column=ii, value='Load Factor (Mode NO.2)')
            #     for i in range(len(LF2)):
            #         sheet1.cell(row=i + 2, column=ii, value=LF2[i])
            #     ii += 1
            # if len(Buckling_data['Factors3']) > 0:
            #     LF3 = np.array(Buckling_data['Factors3'])
            #     sheet1.cell(row=1, column=ii, value='Load Factor (Mode NO.3)')
            #     for i in range(len(LF3)):
            #         sheet1.cell(row=i + 2, column=ii, value=LF3[i])
            #     ii += 1
                # Open a dialog to choose the save location
                save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", str("Global Buckling Sect_"),"Excel Files (*.xlsx)")
        if save_path:
            # Save the workbook to the chosen location
            wb.save(save_path)

    @Slot()
    def on_Close_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError
        QDialog.close(self)

    def find_min_elements(self, list1, list2, list3, list4):
        max_len = max(len(list1), len(list2), len(list3), len(list4))
        result = []
        for i in range(max_len):
            val1 = list1[i] if i < len(list1) else float('inf')
            val2 = list2[i] if i < len(list2) else float('inf')
            val3 = list3[i] if i < len(list3) else float('inf')
            val4 = list4[i] if i < len(list4) else float('inf')
            result.append(min(val1, val2, val3, val4))
        return result

    def solve_cubic(self, list1, list2, list3, r, vs, ws): # (Pv, Pw, Pr, r, vs, ws)
        # Solve a Pcr equation
        roots = []
        max_len = max(len(list1), len(list2), len(list3))
        for i in range(max_len):
            Pv = list1[i] if i < len(list1) else float(0) # Pv
            Pw = list2[i] if i < len(list2) else float(0) # Pw
            Pr = list3[i] if i < len(list3) else float(0) # Pr
            roots.append(np.min(np.roots([(r ** 2 - vs ** 2 - ws ** 2), - ((Pv + Pw + Pr) * r ** 2 - Pw * vs ** 2 - Pv * ws **2),
                                   (r ** 2 * (Pv * Pw + Pw * Pr + Pr * Pv)), - (Pv * Pw * Pr * r ** 2)])))
        return roots


