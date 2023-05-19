# -*- coding: utf-8 -*-

"""
Module implementing ShowResultsYS_Dialog.
"""
import traceback
from PySide6.QtCore import Slot, Qt
from PySide6.QtGui import QDoubleValidator
from PySide6.QtWidgets import QDialog, QHeaderView, QAbstractItemView, QTableWidgetItem
from pyqtgraph import opengl as gl
from .Ui_ShowResultsYieldS import Ui_ShowResultsYS_Dialog
from PySide6.QtWidgets import QApplication, QMainWindow, QSizePolicy, QWidget, QVBoxLayout, QStyledItemDelegate
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
import numpy as np
from openpyxl import Workbook
from PySide6 import QtWidgets
import matplotlib.font_manager as font_manager
from gui.msasect.base.Model import msaModel
from analysis.CMSect.variables.Model import YieldSAnalResults as CMYieldSAnalResults
from analysis.FESect.variables.Model import YieldSAnalResults as FEYieldSAnalResults
from analysis.CMSect.variables.Model import YieldSurfaceAnalInfo as CMYSAnalInfo
from analysis.FESect.variables.Model import YieldSurfaceAnalInfo as FEYSAnalInfo
from analysis.CMSect.util.GetPlanDataByAngle import GetPDataByAng
from analysis.CMSect.util import GetSectionCapacityFactor
from analysis.CMSect.util import GetSectionCapacityFactor2D
# ####Only for test!!!
# import pandas as pd

class NumericDelegate(QStyledItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.validator = QDoubleValidator()

    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        editor.setValidator(self.validator)
        return editor


class ShowResultsYS_Dialog(QDialog, Ui_ShowResultsYS_Dialog):
    """
    Class documentation goes here.
    """

    def __init__(self, mw, parent):
        """
        Constructor

        @param parent reference to the parent widget (defaults to None)
        @type QWidget (optional)
        """
        super().__init__(parent=parent)
        self.setupUi(self)
        self.mw = mw
        self.parent_window = self.parent()
        self.grandparent_window = self.parent_window.parent()
        self.PeY = np.array([])
        self.PeX = np.array([])
        self.PeZ = np.array([])
        self.Plan2DPts = np.array([])
        self.TableHX = ''
        self.TableHY = ''
        self.TableHZ = ''

        if self.grandparent_window.Centerline_radioButton.isChecked():
            if len(CMYieldSAnalResults.ONx) > 0 and len(CMYieldSAnalResults.OMy) > 0 and len(CMYieldSAnalResults.OMz) > 0:
                self.parent_window.ShowResults_pushButton.setEnabled(True)
        elif self.grandparent_window.Outline_radioButton.isChecked():
            if (len(FEYieldSAnalResults.ONx) > 0 and len(FEYieldSAnalResults.OMy) > 0 and len(FEYieldSAnalResults.OMz) > 0):
                self.parent_window.ShowResults_pushButton.setEnabled(True)

        # if (len(YieldSAnalResults.ONx) > 0 and len(YieldSAnalResults.OMy) > 0 and len(YieldSAnalResults.OMz) > 0) or \
        #         (len(FEYieldSAnalResults.ONx) > 0 and len(FEYieldSAnalResults.OMy) > 0 and len(FEYieldSAnalResults.OMz) > 0):
        #     self.mw.ShowResults_pushButton.setEnabled(True)
        # elif len(YieldSAnalResults.ONx_y) > 0 and len(YieldSAnalResults.OMy_x) > 0:
        #     self.mw.PYSPMy_radioButton.setEnabled(True)
        ##
        # grid = gl.GLGridItem()
        # grid.scale(2, 2, 1)
        # grid.setDepthValue(10)
        # self.FullYS_graphicsView.addItem(grid)
        # ax = gl.GLAxisItem()
        # ax.setSize(40, 40, 40)
        # self.FullYS_graphicsView.addItem(ax)
        # self.FullYS_graphicsView.addItem(grid)
        # self.FullYS_graphicsView.addItem(ax)
        ##
        # Create a figure canvas to display the 3D plot
        self.fig3D = plt.figure(facecolor="black")
        self.canvas = FigureCanvas(self.fig3D)
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.FYSShow_verticalLayout.addWidget(self.canvas)
        ##
        self.setFixedSize(1350, 705)
        ##
        self.fig2D = Figure(facecolor="black")
        self.ax = self.fig2D.add_subplot(111)
        self.canvas2D = FigureCanvas(self.fig2D)
        self.YS2DPlot_verticalLayout_2.addWidget(self.canvas2D)
        ##
        self.lines = []
        ##
        self.LPTableData_dict = {}
        ##
        self.LoadingP_tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.LoadingP_tableWidget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.LoadingP_tableWidget.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.LoadingP_tableWidget.verticalHeader().setVisible(False)
        self.LoadingP_tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.LoadingP_tableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
        # self.LoadingP_tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.LoadingP_tableWidget.horizontalHeader().setSectionsClickable(False)
        self.LoadingP_tableWidget.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.LoadingP_tableWidget.verticalHeader().setDefaultAlignment(Qt.AlignCenter)

        # Connect cellChanged signal to custom slot
        self.LoadingP_tableWidget.cellChanged.connect(self.on_cell_changed)
        self.LoadingP_tableWidget.cellClicked.connect(self.on_cell_clicked)
        numeric_delegate = NumericDelegate(self.LoadingP_tableWidget)
        self.LoadingP_tableWidget.setItemDelegate(numeric_delegate)
        ##
        self.LP_scatter = None
        ##
        self.state_ShowPoints = 1
        self.state_ShowLines = 1
        self.ShowPoints_checkBox.stateChanged.connect(
            lambda state: self.Checkbox_on_state_changed(self.ShowPoints_checkBox, state))
        self.ShowLines_checkBox.stateChanged.connect(
            lambda state: self.Checkbox_on_state_changed(self.ShowLines_checkBox, state))
        # self.UBorder = self.FullYS_graphicsView.addPlot(row=0, col=0)
        # Create the 3D plot
        # self.updatTableStatus()
        self.initDialog()
        self.initLoadingPTable()

    def initDialog(self):
        ##
        self.updatTableStatus()
        #self.setFixedSize(self.width(), self.height())
        self.setFixedSize(self.size())
        ##
        self.ax3D = self.fig3D.add_subplot(projection='3d')
        self.YieldS3DPlot()
        self.YieldS2DPlot(1.0, self.Plan2DPts)

    # def updat_table_status(self):
    #     row_count = self.LoadingP_tableWidget.rowCount()
    #     column_count = self.LoadingP_tableWidget.columnCount()
    #     for i in range(row_count):
    #         item = QTableWidgetItem()
    #         item.setText(str(i + 1))
    #         self.LoadingP_tableWidget.setItem(i+1, 0, item)
    #     ##
    #     for j in range(column_count):
    #         item = self.LoadingP_tableWidget.item(0, j)
    #         item.setFlags(item.flags() ^ Qt.ItemIsEditable ^ Qt.ItemIsSelectable)
    #     for ii in range(row_count):
    #         item = self.LoadingP_tableWidget.item(ii, 0)
    #         item.setFlags(item.flags() ^ Qt.ItemIsEditable ^ Qt.ItemIsSelectable)
    #     ##
    #     # Center contents of All cell
    #     for row in range(row_count):
    #         for column in range(row_count):
    #             item = self.LoadingP_tableWidget.item(row, column)
    #             if item is not None:
    #                 item.setTextAlignment(Qt.AlignCenter)

    def initLoadingPTable(self):
        ##
        self.LoadingP_tableWidget.setRowCount(2)
        # LP2DTable = self.LoadingP_2D_tableWidget()
        for row in range(2):
            for col in range(self.LoadingP_tableWidget.columnCount()):
                item = QTableWidgetItem()
                item.setTextAlignment(Qt.AlignCenter)
                self.LoadingP_tableWidget.setItem(row, col, item)
        ##
        self.LoadingP_tableWidget.resizeColumnsToContents()
        # self.LoadingP_tableWidget.setSpan(0, 0, 1, 0)  ## ID
        ##
        # self.LoadingP_tableWidget.setHorizontalHeaderLabels(["ID", "Px", "Mv", "Mw", "SCF"])
        #
        # # for col, text in enumerate(['', 'kN', 'kN.m', 'kN.m', '']):
        # for col, text in enumerate(['', '', '', '', '']):
        #     item = QTableWidgetItem(text)
        #     item.setTextAlignment(Qt.AlignCenter)
        #     self.LoadingP_tableWidget.setItem(0, col, item)
        # self.LoadingP_tableWidget.setSpan(0, 4, 1, 4)  ## SCF
        # Set column headings
        if self.parent_window.PrincipalAxis_radioButton.isChecked():
            column_labels = ["ID", "Px", "Mv", "Mw", "SCF"]
        else:
            column_labels = ["ID", "Px", "My", "Mz", "SCF"]
        self.LoadingP_tableWidget.setHorizontalHeaderLabels(column_labels)
        self.LoadingP_tableWidget.horizontalHeader().setVisible(True)

        # Move the table header text to the cell in the first row and set the text to be centered
        # for i, label in enumerate(column_labels):
        #     item = QTableWidgetItem(label)
        #     item.setTextAlignment(Qt.AlignCenter)
        #     item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Set the non-editable flag
        #     self.LoadingP_tableWidget.setItem(0, i, item)
        #
        ##
        row_count = self.LoadingP_tableWidget.rowCount()
        for i in range(row_count):
            item = QTableWidgetItem()
            item.setText(str(i + 1))
            # self.LoadingP_tableWidget.setItem(i + 1, 0, item)
            self.LoadingP_tableWidget.setItem(i, 0, item)
        ##
        # Center contents of All cell
        for row in range(row_count):
            for column in range(row_count):
                item = self.LoadingP_tableWidget.item(row, column)
                if item is not None:
                    item.setTextAlignment(Qt.AlignCenter)
        # for j in range(self.LoadingP_tableWidget.columnCount()):
        #     item = self.LoadingP_tableWidget.item(0, j)
        #     item.setFlags(item.flags() ^ Qt.ItemIsEditable ^ Qt.ItemIsSelectable)
        for ii in range(self.LoadingP_tableWidget.rowCount()):
            item = self.LoadingP_tableWidget.item(ii, 0)
            item.setFlags(item.flags() ^ Qt.ItemIsEditable ^ Qt.ItemIsSelectable)
        ##
        # Set the first column adaptive width, Set other column equal widths
        header = self.LoadingP_tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        for col in range(1, self.LoadingP_tableWidget.columnCount()):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)


    def YieldS3DPlot(self):
        ##for testing
        ## Saddle example with x and y specified
        if self.grandparent_window.Centerline_radioButton.isChecked():
            YSAnalResults = CMYieldSAnalResults
            MStep = CMYSAnalInfo.MStep
        elif self.grandparent_window.Outline_radioButton.isChecked():
            YSAnalResults = FEYieldSAnalResults
            MStep = FEYSAnalInfo.MStep
        ploteZ = YSAnalResults.ONx
        ploteY = YSAnalResults.OMy
        ploteX = YSAnalResults.OMz
        tploteZ = ploteZ.values()
        tploteX = ploteX.values()
        tploteY = ploteY.values()
        ##
        self.PeZ = np.array(list(tploteZ))
        self.PeX = np.array(list(tploteX))
        self.PeY = np.array(list(tploteY))

        # x = np.random.uniform(-1, 1, 100)
        # y = np.random.uniform(-1, 1, 100)
        # z = np.random.uniform(-1, 1, 100)
        ###-------------------------------------------------------
        # 假设 self.PeX, self.PeY, self.PeZ 是你的数据
        PeX = np.array(self.PeX)
        PeY = np.array(self.PeY)
        PeZ = np.array(self.PeZ)

        # 每三个元素之间插入一个NaN值
        # PeX_nan = np.insert(PeX, np.arange(3, len(PeX), 3), np.nan)
        # PeY_nan = np.insert(PeY, np.arange(3, len(PeY), 3), np.nan)
        # PeZ_nan = np.insert(PeZ, np.arange(3, len(PeZ), 3), np.nan)
        # 计算组的数量
        num_groups = len(PeZ) // int(MStep+1)

        # 重新整理数据，使每一组数据成为一个单独的实体
        PeX_grouped = PeX.reshape(num_groups, int(MStep+1))
        PeY_grouped = PeY.reshape(num_groups, int(MStep+1))
        PeZ_grouped = PeZ.reshape(num_groups, int(MStep+1))
        ###-------------------------------------------------------

        # Plot the data
        # ax.scatter(self.PeX, self.PeY, self.PeZ, s=5, alpha=0.6, c=self.PeZ, cmap='jet')
        # for i in range(len(self.PeX) - 1):
        #     ax.plot(self.PeX[i:i + 2], self.PeY[i:i + 2], self.PeZ[i:i + 2], color='r', linestyle='-', linewidth=2)
        if self.state_ShowPoints == 1 and self.state_ShowLines == 0:
            #self.ax3D.scatter(self.PeX, self.PeY, self.PeZ, s=5, alpha=0.6, c=self.PeZ, cmap='jet')
            self.ax3D.clear()
            self.ax3D.plot(self.PeX, self.PeY, self.PeZ, marker='o', linestyle='', color='r', markersize=3)
        elif self.state_ShowPoints == 0 and self.state_ShowLines == 1:
            self.ax3D.clear()
            # self.ax3D.plot(self.PeX, self.PeY, self.PeZ, marker='o', linestyle='-', color='r', markersize=0)
            for i in range(num_groups):
                self.ax3D.plot(PeX_grouped[i], PeY_grouped[i], PeZ_grouped[i], marker='o', linestyle='-', color='r',
                               markersize=0)
        elif self.state_ShowPoints == 1 and self.state_ShowLines == 1:
            self.ax3D.clear()
            # self.ax3D.plot(self.PeX, self.PeY, self.PeZ, marker='o', linestyle='-', color='r', markersize=3)
            # self.ax3D.plot(PeX_nan, PeY_nan, PeZ_nan, marker='o', linestyle='-', color='r', markersize=3)
            for i in range(num_groups):
                self.ax3D.plot(PeX_grouped[i], PeY_grouped[i], PeZ_grouped[i], marker='o', linestyle='-', color='r',
                               markersize=3)
        elif self.state_ShowPoints == 0 and self.state_ShowLines == 0:
            self.ax3D.clear()
            # self.ax3D.plot(self.PeX, self.PeY, self.PeZ, marker='o', linestyle='-', color='black', markersize=0)
            for i in range(num_groups):
                self.ax3D.plot(PeX_grouped[i], PeY_grouped[i], PeZ_grouped[i], marker='o', linestyle='', color='r',
                               markersize=0)
        self.fig3D.canvas.draw_idle()
        # ax.plot(self.PeX, self.PeY, self.PeZ, marker='o', linestyle='-', color='r', markersize=3)
        self.ax3D.set_facecolor("black")
        font = {'family': 'Times New Roman', 'size': 10, 'weight': 'bold'}
        # Set the axis labels
        # plt.title("This is a complete view of the yield surface!")
        if self.parent_window.PrincipalAxis_radioButton.isChecked():
            self.ax3D.set_xlabel('Bending Moment about w-axis Mw', color='white', fontsize=10, fontdict=font)
            self.ax3D.set_ylabel('Bending Moment about v-axis Mv', color='white', fontsize=10, fontdict=font)
            self.ax3D.set_zlabel('Axial Force Px', color='white', fontsize=10, fontdict=font)
            self.TableHX = "Mw"
            self.TableHY = "Mv"
            self.TableHZ = "Px"
            plt.title("Px vs Mv vs Mw", color='white', fontdict=font)
        else:
            self.ax3D.set_xlabel('Bending Moment about z-axis Mz', color='white', fontsize=10, fontdict=font)
            self.ax3D.set_ylabel('Bending Moment about y-axis My', color='white', fontsize=10, fontdict=font)
            self.ax3D.set_zlabel('Axial Force Px', color='white', fontsize=10, fontdict=font)
            self.TableHX = "Mz"
            self.TableHY = "My"
            self.TableHZ = "Px"
            plt.title("Px vs My vs Mz", color='white', fontdict=font)
        ## Adjust this value to increase/decrease distance between label and axis
        self.ax3D.xaxis.labelpad = 10
        self.ax3D.yaxis.labelpad = 10
        self.ax3D.zaxis.labelpad = 16

        labelpad = 1  # Adjust this value to increase/decrease distance between label and axis
        plt.tight_layout(pad=labelpad)
        # self.ax3D.ticklabel_format(style='plain', axis='x', scilimits=(0, 0))
        # self.ax3D.ticklabel_format(style='plain', axis='y', scilimits=(0, 0))
        # self.ax3D.ticklabel_format(style='plain', axis='z', scilimits=(0, 0))
        self.ax3D.grid(False)
        ##
        self.ax3D.w_xaxis.line.set_color('white')
        self.ax3D.w_yaxis.line.set_color('white')
        self.ax3D.w_zaxis.line.set_color('white')

        self.ax3D.xaxis.pane.set_edgecolor('white')
        self.ax3D.yaxis.pane.set_edgecolor('white')
        self.ax3D.zaxis.pane.set_edgecolor('white')
        # ax.xaxis.pane.set_edgecolor((0, 0, 0, 0))
        # ax.yaxis.pane.set_edgecolor((0, 0, 0, 0))
        # ax.zaxis.pane.set_edgecolor((0, 0, 0, 0))

        self.ax3D.xaxis.pane.set_facecolor('black')
        self.ax3D.yaxis.pane.set_facecolor('black')
        self.ax3D.zaxis.pane.set_facecolor('black')

        self.ax3D.xaxis.label.set_color('white')
        self.ax3D.yaxis.label.set_color('white')
        self.ax3D.zaxis.label.set_color('white')

        self.ax3D.tick_params(axis='x', colors='white', pad=2)
        self.ax3D.tick_params(axis='y', colors='white', pad=2)
        self.ax3D.tick_params(axis='z', colors='white', pad=10)
        ##
        self.ax3D.w_xaxis.line.set_linewidth(1.2)
        self.ax3D.w_yaxis.line.set_linewidth(1.2)
        self.ax3D.w_zaxis.line.set_linewidth(1.2)
        ##
        self.ax3D.xaxis.set_tick_params(width=2, color='white', length=5)
        self.ax3D.yaxis.set_tick_params(width=2, color='white', length=5)
        self.ax3D.zaxis.set_tick_params(width=2, color='white', length=5)
        ##
        self.ax3D.view_init(elev=30, azim=45)
        ##
        self.ax3D.xaxis.set_major_formatter(plt.FormatStrFormatter('%.1e'))
        self.ax3D.yaxis.set_major_formatter(plt.FormatStrFormatter('%.1e'))
        self.ax3D.zaxis.set_major_formatter(plt.FormatStrFormatter('%.1e'))
        ##
        for label in self.ax3D.get_xticklabels() + self.ax3D.get_yticklabels() + self.ax3D.get_zticklabels():
            label.set_fontname('Times New Roman')
        ##
        # plt.show()
        # x = np.linspace(-8, 8, 50)
        # y = np.linspace(-8, 8, 50)
        # z = 0.1 * ((x.reshape(50, 1) ** 2) - (y.reshape(1, 50) ** 2))
        # print("PeZ=", PeZ)
        # # print("PeX=", PeX)
        # # print("PeY=", PeY)
        # print("ploteZ=", ploteZ)
        # print("the length of PeZ=", len(PeZ))
        # print("the length of PeX=", len(PeX))
        # print("the length of PeY=", len(PeY))
        # print("ploteX=", ploteX)
        # print("ploteY=", ploteY)
        # Create the plot canvas
        # self.canvas = FigureCanvas(plt.Figure(figsize=(5, 4), dpi=100))

        # Add the plot canvas to the layout


        # Create the 3D plot
        # ax = self.canvas.figure.add_subplot(111, projection='3d')
        # x = np.linspace(-10, 10, 100)
        # y = np.linspace(-10, 10, 100)
        # X, Y = np.meshgrid(x, y)
        # Z = np.sin(np.sqrt(X ** 2 + Y ** 2))
        #
        # # Plot the surface
        # ax.plot_surface(X, Y, Z)


        # fig = plt.figure()
        # ax1 = plt.axes(projection='3d')
        # ax1.scatter3D(PeX, PeY, PeZ, cmap='Blues')  # 绘制散点图
        # # ax1.plot3D(x, y, z, 'gray')  # 绘制空间曲线
        # plt.show()
        # # PlotYS3D = gl.GLSurfacePlotItem(x=PeX, y=PeY, z=PeZ, shader='normalColor')
        # # PlotYS3D.translate(-10, -10, 0)
        # self.FullYS_graphicsView.addItem(ax1)
    def Updat2Dplot_Scatters(self, row, tPx, tMy, tMz, Plan2DPts, tcolor):
        ##

        ##
        # scatter_points = self.ax.collections
        # if row in self.LPTableData_dict.keys():
        #     tPx1 = float(self.LPTableData_dict[row][0])
        #     tMy1 = float(self.LPTableData_dict[row][1])
        #     tMz1 = float(self.LPTableData_dict[row][2])
        #     if tMz1 >= 0.0:
        #         tSqrtM = np.sqrt(tMy1**2+tMz1**2)
        #     else:
        #         tSqrtM = -np.sqrt(tMy1 ** 2 + tMz1 ** 2)
        #     #
        #     for scatter in scatter_points:
        #         x_data, y_data = scatter.get_offsets().T
        #         index = np.where((x_data == tSqrtM) & (y_data == tPx1))[0]
        #         if len(index) > 0:
        #             scatter.remove()
            # for (tx1, ty1) in scatter_points:
            #     (tx1, ty1).remove()

        self.LPTableData_dict[row] = tuple([tPx, tMy, tMz, Plan2DPts, tcolor])
        ##
        self.YieldS2DPlot(tPx, Plan2DPts)
        ##
        self.ScatterPlot(tPx, tMy, tMz, tcolor)
        ##
        self.adjust_axes()
        # if tMz >= 0.0:
        #     tx_value = np.sqrt(tMy ** 2 + tMz ** 2)
        # else:
        #     tx_value = -np.sqrt(tMy ** 2 + tMz ** 2)
        # ##
        # self.LP_scatter = self.ax.scatter(tx_value, tPx, color=tcolor, marker='o')
        ##
        # self.ax.autoscale(enable=True, axis='both', tight=False)
        # self.ax.relim()
        # self.ax.autoscale_view()
        #
        self.canvas2D.draw()

    def ScatterPlot(self, tPx, tMy, tMz, tcolor):
        ##
        for collection in self.ax.collections:
            if collection == self.LP_scatter:
                collection.remove()
        # if self.LP_scatter is not None:
        #     self.LP_scatter.remove()
        #
        if tMz >= 0.0:
            tx_value = np.sqrt(tMy ** 2 + tMz ** 2)
        else:
            tx_value = -np.sqrt(tMy ** 2 + tMz ** 2)
        ##
        self.LP_scatter = self.ax.scatter(tx_value, tPx, color=tcolor, marker='o')
        plt.draw()
        return

    def adjust_axes(self):
        ##
        # Get the coordinates of the point created by the plot function
        x_line_coords, y_line_coords = self.Plot2D_curve.get_data()
        # Get the coordinates of the point created by the scatter function
        scatter_points = self.ax.collections
        coords = np.vstack([point.get_offsets() for point in scatter_points])
        x_scatter_coords, y_scatter_coords = coords.T
        # Merge x-axis and y-axis coordinates
        x_coords = np.concatenate((x_line_coords, x_scatter_coords))
        y_coords = np.concatenate((y_line_coords, y_scatter_coords))
        #
        ## Get the maximum value of x-axis and y-axis
        max_x = max(x_coords)
        max_y = max(y_coords)
        ## Get the minimum value of x-axis and y-axis
        min_x = min(x_coords)
        min_y = min(y_coords)
        ## setting the current
        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()
        # # Adjust the axis range according to the new point
        # new_xlim = (min(xlim[0], max_x)-max_x*0.1, max(xlim[1], max_x)+max_x*0.1)
        # new_ylim = (min(ylim[0], max_y)-max_y*0.1, max(ylim[1], max_y)+max_y*0.1)
        new_xlim = (min(min_x, max_x)-max_x*0.1, max(xlim[1], max_x)+max_x*0.1)
        new_ylim = (min(min_y, max_y)-max_y*0.1, max(ylim[1], max_y)+max_y*0.1)
        # Update axis range
        self.ax.set_xlim(new_xlim)
        self.ax.set_ylim(new_ylim)
        ##

    def add_line(self, x, y):
        # x = np.array([1, 2, 3, 4, 5])
        # y = np.array([2, 4, 6, 8, 10])
        self.LP_line, = self.ax.plot(x, y, color='white', linestyle='-', linewidth=0.8, label="LP_line")
        self.lines.append(self.LP_line)
        self.canvas.draw()

    def updatTableStatus(self):
        row_count = self.LoadingP_tableWidget.rowCount()
        column_count = self.LoadingP_tableWidget.columnCount()
        for i in range(row_count):
            item = QTableWidgetItem()
            item.setText(str(i + 1))
            # self.LoadingP_tableWidget.setItem(i+1, 0, item)
            self.LoadingP_tableWidget.setItem(i, 0, item)
        ##
        # for j in range(column_count):
        #     item = self.LoadingP_tableWidget.item(0, j)
        #     if item is not None:
        #         item.setFlags(item.flags() ^ Qt.ItemIsEditable ^ Qt.ItemIsSelectable)
        for ii in range(row_count):
            item = self.LoadingP_tableWidget.item(ii, 0)
            if item is not None:
                item.setFlags(item.flags() ^ Qt.ItemIsEditable ^ Qt.ItemIsSelectable)
        ##
        # Center contents of All cell
        for row in range(row_count):
            for column in range(row_count):
                item = self.LoadingP_tableWidget.item(row, column)
                if item is not None:
                    item.setTextAlignment(Qt.AlignCenter)
        ##
        # validator = QDoubleValidator()
        # for i in range(self.LoadingP_tableWidget.rowCount()):
        #     for j in range(self.LoadingP_tableWidget.columnCount()):
        #         item = QTableWidgetItem()
        #         item.setFlags(item.flags() | Qt.ItemIsEditable)
        #         # item.setData(Qt.EditRole, 0)  # 设置初始值为0
        #         self.LoadingP_tableWidget.setItem(i, j, item)
        #         self.LoadingP_tableWidget.item(i, j).setValidator(validator)

    def on_cell_clicked(self, row, column):
        if row == self.LoadingP_tableWidget.rowCount() - 1:
            self.LoadingP_tableWidget.insertRow(self.LoadingP_tableWidget.rowCount())
        ##
        self.updatTableStatus()
        ##
        if row in self.LPTableData_dict.keys():
            tPx = float(self.LPTableData_dict[row][0])
            tMy = float(self.LPTableData_dict[row][1])
            tMz = float(self.LPTableData_dict[row][2])
            tPlan2Dplots = self.LPTableData_dict[row][3]
            tcolor = self.LPTableData_dict[row][4]
            #
            self.YieldS2DPlot(tPx, tPlan2Dplots)
            ##
            self.ScatterPlot(tPx, tMy, tMz, tcolor)
            ##
            self.adjust_axes()
            #
            self.canvas2D.draw()
        else:
            if self.Plot2D_curve in self.ax.lines:
                try:
                    self.Plot2D_curve.remove()
                    self.canvas2D.draw_idle()
                except AttributeError:
                    traceback.print_exc()
            # if self.Plot2D_curve in self.ax.lines:
            #     try:
            #         self.ax.lines.remove(self.Plot2D_curve)
            #     except AttributeError:
            #         traceback.print_exc()
            # self.ax.clear()
            ##
            for collection in self.ax.collections:
                if collection == self.LP_scatter:
                    collection.remove()
            ##
            self.canvas2D.draw()
            # if self.LP_scatter is not None:
            #     self.LP_scatter.remove()


    def on_cell_changed(self, row, column):
        # if row == 0:
        #     return
        ##
        if column not in [1, 2, 3]:
            return
        ##
        item1 = self.LoadingP_tableWidget.item(row, 1)
        item2 = self.LoadingP_tableWidget.item(row, 2)
        item3 = self.LoadingP_tableWidget.item(row, 3)
        if item1 is not None and item2 is not None and item3 is not None:
            try:
                tPx = float(item1.text())  # if item1 is not None else 0 value1
                tMy = float(item2.text())  # if item2 is not None else 0 value2
                tMz = float(item3.text())  # if item3 is not None else 0 value3
                ##
                listSpatialpoints = np.concatenate((self.PeX.reshape(-1, 1), self.PeY.reshape(-1, 1), self.PeZ.reshape(-1, 1)), axis=1)
                ## Only for testing!!!!!!!!!!!!!!!!!!!!!!!!!
                # YSData = pd.DataFrame(listSpatialpoints)
                # writer = pd.ExcelWriter('Test Yield Surface03.xlsx')
                # YSData.to_excel(writer, 'page_1', float_format='%.5f')
                # writer.save()
                # writer.close()
                ## !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
                if self.grandparent_window.Centerline_radioButton.isChecked():
                    tFlag = 1 ## For CMSect
                elif self.grandparent_window.Outline_radioButton.isChecked():
                    tFlag = 2  ## For FESect
                Plan2DPts, Plan3DPts = GetPDataByAng(listSpatialpoints, tMy, tMz, tPx, tFlag)
                ## Only for testing!!!!!!!!!!!!!!!!!!!!!!!!!
                # YSData01 = pd.DataFrame(Plan2DPts)
                # YSData02 = pd.DataFrame(Plan3DPts)
                # with pd.ExcelWriter('C:/Users/gaowl/PycharmProjects/Mastan3/Source/gui/msasect/examples/TestPost3D_PlanarData.xlsx') as writer:
                #     # 将 DataFrame 写入 Excel 文件
                #     YSData01.to_excel(writer, sheet_name='Sheet1', index=False)
                # with pd.ExcelWriter('C:/Users/gaowl/PycharmProjects/Mastan3/Source/gui/msasect/examples/TestPost3D_3DData.xlsx') as writer:
                #     # 将 DataFrame 写入 Excel 文件
                #     YSData02.to_excel(writer, sheet_name='Sheet1', index=False)
                ## !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
                LP_x = np.sqrt(tMy**2+tMz**2)
                LP_y = tPx
                if abs(LP_x) < 1e-3 and abs(LP_y) < 1e-3:
                    SCF = 0.0
                else:
                    SCF = GetSectionCapacityFactor2D.GetSCF2D(Plan2DPts, np.array([LP_x, LP_y]))
                ##
                if SCF > 1.0:
                    tcolor = 'red'
                else:
                    tcolor = 'green'
                self.Updat2Dplot_Scatters(row, tPx, tMy, tMz, Plan2DPts, tcolor)
                item = QTableWidgetItem(str(format(SCF, '.2f')))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.LoadingP_tableWidget.setItem(row, 4, item)
                print(f"Row {row}, Col 1: {tPx}, Col 2: {tMy}, Col 3: {tMz}")
                # print("Col 1:", value1)
                # print("Col 2:", value2)
            except ValueError:
                #traceback.print_exc()
                print()

    def YieldS2DPlot(self, Px, Plan2DPts):
        ## Plot Setting
        self.ax.spines['bottom'].set_linewidth(1.6)
        self.ax.spines['left'].set_linewidth(1.6)
        ##
        #self.ax.set_facecolor('#FFFFE0')
        ##
        # self.fig2D.tight_layout(pad=2.0, h_pad=None, w_pad=None, rect=None)
        # Set the labels and title
        self.ax.tick_params(direction='in', axis='both', which='major', labelsize=9, pad=6)
        self.ax.set_facecolor("black")
        self.ax.tick_params(axis="x", colors="white", labelsize=9)
        self.ax.tick_params(axis="y", colors="white", labelsize=9)
        self.ax.spines['bottom'].set_color('white')
        self.ax.spines['top'].set_color('white')
        self.ax.spines['right'].set_color('white')
        self.ax.spines['left'].set_color('white')
        # get font properties and set size
        font = font_manager.FontProperties()
        font.set_size(self.fig2D.dpi * 0.06)
        # set font properties for x and y-axis labels
        self.ax.xaxis.label.set_font_properties(font)
        self.ax.yaxis.label.set_font_properties(font)
        # Set the coordinate axis number format to fixed-point notation.
        # self.ax.xaxis.set_major_formatter(plt.FormatStrFormatter('%.2f'))
        # self.ax.yaxis.set_major_formatter(plt.FormatStrFormatter('%.2f'))
        # Create the data
        if Plan2DPts.size == 0:
            # x = [1, 2, 3, 4, 5]
            # y = [1, 4, 9, 16, 25]
            x = []
            y = []
            self.ax.set_xlim(-1, 1)
            self.ax.set_ylim(-1, 1)
            self.fig2D.tight_layout(pad=1.0, h_pad=None, w_pad=None, rect=(0.1, 0.05, 0.95, 0.95))
            self.ax.xaxis.set_major_formatter(plt.FormatStrFormatter('%.1f'))
            self.ax.yaxis.set_major_formatter(plt.FormatStrFormatter('%.1f'))
        else:
            self.ax.clear()
            x = Plan2DPts[:, 0].tolist()
            y = Plan2DPts[:, 1].tolist()
            self.fig2D.tight_layout(pad=1.0, h_pad=None, w_pad=None, rect=(0.1, 0.05, 0.95, 0.95))
        # Create the subplot
        #ax = self.fig2D.add_subplot(111)
        # self.fig2D.tight_layout(pad=1.0, h_pad=None, w_pad=None, rect=None)
        # Plot the data
        #self.ax.plot(x, y)      ## type(x)=type(y)=list, like x = [1, 2, 3, 4, 5]
        self.Plot2D_curve, = self.ax.plot(x, y, label="Plot2D_curve", color="red") ## type(x)=type(y)=list, like x = [1, 2, 3, 4, 5]
        font = {'family': 'Times New Roman', 'size': 12, 'weight': 'bold'}
        # Set the labels and title
        if abs(Px) < 1e-3:
            if self.parent_window.PrincipalAxis_radioButton.isChecked():
                self.ax.set_xlabel('Bending Moment Mw', color='white', fontsize=10, fontdict=font)
                self.ax.set_ylabel('Bending Moment Mv', color='white', fontsize=10, fontdict=font)
            else:
                self.ax.set_xlabel('Bending Moment Mz', color='white', fontsize=10, fontdict=font)
                self.ax.set_ylabel('Bending Moment My', color='white', fontsize=10, fontdict=font)
        else:
            self.ax.set_xlabel('Bending Moment M', color='white', fontsize=10, fontdict=font)
            self.ax.set_ylabel('Axial Force Px', color='white', fontsize=10, fontdict=font)
        self.ax.set_title('Moment Interaction Curve', fontsize=10, color='white', fontdict=font)
        self.ax.grid(linestyle=':', linewidth='0.8', color='gray', alpha=0.8)
        ##

        # self.ax.spines['bottom'].set_linewidth(1.6)
        # self.ax.spines['left'].set_linewidth(1.6)
        # ##
        # #self.ax.set_facecolor('#FFFFE0')
        # ##
        # self.fig2D.tight_layout(pad=1.08, h_pad=None, w_pad=None, rect=None)
        # # Set the labels and title
        # self.ax.tick_params(direction='in', axis='both', which='major', labelsize=9, pad=6)
        # self.ax.set_facecolor("black")
        # self.ax.tick_params(axis="x", colors="white", labelsize=9)
        # self.ax.tick_params(axis="y", colors="white", labelsize=9)
        # self.ax.spines['bottom'].set_color('white')
        # self.ax.spines['top'].set_color('white')
        # self.ax.spines['right'].set_color('white')
        # self.ax.spines['left'].set_color('white')
        #plt.rcParams['axes.labelsize'] = 'large'
        # plt.tick_params(axis='both', which='major', labelsize=12, pad=6)
        # plt.tight_layout()

        # # get font properties and set size
        # font = font_manager.FontProperties()
        # font.set_size(self.fig2D.dpi * 0.06)
        # # set font properties for x and y-axis labels
        # self.ax.xaxis.label.set_font_properties(font)
        # self.ax.yaxis.label.set_font_properties(font)
        # Set the coordinate axis number format to fixed-point notation.
        self.ax.xaxis.set_major_formatter(plt.FormatStrFormatter('%.1e'))
        self.ax.yaxis.set_major_formatter(plt.FormatStrFormatter('%.1e'))
        for label in self.ax.get_xticklabels() + self.ax.get_yticklabels():
            label.set_fontname('Times New Roman')
        ## Update canvas
        self.canvas2D.draw()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.fig3D.set_size_inches(event.size().width() / self.fig3D.dpi, event.size().height() / self.fig3D.dpi)
        self.fig2D.set_size_inches(event.size().width() / self.fig2D.dpi, event.size().height() / self.fig2D.dpi)

    @Slot()
    def on_ClearInput_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError
        scatter_points = self.ax.collections
        for scatter in scatter_points:
            scatter.remove()
        ##
        self.ax.autoscale_view()
        self.ax.autoscale(enable=True, axis='both', tight=False)
        self.ax.relim()
        #
        self.canvas.draw()
        ##
        self.initLoadingPTable()
        #
        for row in range(1, self.LoadingP_tableWidget.rowCount()):
            for col in range(1, self.LoadingP_tableWidget.columnCount()):
                item = self.LoadingP_tableWidget.item(row, col)
                if item is not None:
                    self.LoadingP_tableWidget.setItem(row, col, None)
        ##
        self.LPTableData_dict = {}
        ##
        if self.Plot2D_curve in self.ax.lines:
            try:
                self.Plot2D_curve.remove()
                self.canvas2D.draw_idle()
            except AttributeError:
                traceback.print_exc()
        ##
        for collection in self.ax.collections:
            if collection == self.LP_scatter:
                collection.remove()
        ##
        self.canvas2D.draw()
        #
        self.updatTableStatus()

    @Slot()
    def on_ExportSP_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError
        wb = Workbook()
        ws = wb.active
        ws.title = "Full"

        # Write the data to the worksheet
        ws.cell(row=1, column=1, value=self.TableHX)
        ws.cell(row=1, column=2, value=self.TableHY)
        ws.cell(row=1, column=3, value=self.TableHZ)
        for i in range(len(self.PeX)):
            ws.cell(row=i + 2, column=1, value=self.PeX[i])
            ws.cell(row=i + 2, column=2, value=self.PeY[i])
            ws.cell(row=i + 2, column=3, value=self.PeZ[i])

        # Open a dialog to choose the save location
        save_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")

        if save_path:
            # Save the workbook to the chosen location
            wb.save(save_path)

    @Slot()
    def on_Close_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError
        QDialog.close(self)

    @Slot()
    def on_Delete_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError
        currentRow = self.LoadingP_tableWidget.currentRow()
        if currentRow != -1 and currentRow != 0:
            self.LoadingP_tableWidget.removeRow(currentRow)
            self.LoadingP_tableWidget.clearSelection()
            ##
            if currentRow in self.LPTableData_dict:
                # 删除键及其对应的值
                del self.LPTableData_dict[currentRow]

    @Slot()
    def on_ShowLines_checkBox_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError

    @Slot()
    def on_ShowPoints_checkBox_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        # raise NotImplementedError

    def Checkbox_on_state_changed(self, checkbox, state):
        checkbox_name = checkbox.text()
        if state == 2:  # Checkbox is checked
            if checkbox_name == "Show Points":
                self.state_ShowPoints = 1
            elif checkbox_name == "Show Lines":
                self.state_ShowLines = 1
            # print(f"{checkbox_name} is checked.")
        elif state == 0:  # Checkbox is unchecked
            if checkbox_name == "Show Points":
                self.state_ShowPoints = 0
            elif checkbox_name == "Show Lines":
                self.state_ShowLines = 0
        ##
        # ax = self.fig3D.add_subplot(projection='3d')
        self.YieldS3DPlot()
        # plt.show()
            # print(f"{checkbox_name} is not checked.")
        # elif state == 1:  # Checkbox is partially checked (only for tristate checkboxes)
        #      print(f"{checkbox_name} is partially checked.")